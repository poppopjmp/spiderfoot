# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------

import csv
import html
import json
import logging
import multiprocessing as mp
import openpyxl
import random
import re
import string
import time
import os
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

import cherrypy
import secure
from cherrypy import _cperror
from mako.lookup import TemplateLookup
from mako.template import Template

import markdown

from sflib import SpiderFoot
from sfscan import startSpiderFootScanner
from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.logger import logListenerSetup, logWorkerSetup
from spiderfoot.workspace import SpiderFootWorkspace
from spiderfoot.event_type_mapping import translate_event_type

mp.set_start_method("spawn", force=True)


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(
                f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
                .default_src("'self'")
                .script_src("'self'", "'unsafe-inline'", "blob:")
                .style_src("'self'", "'unsafe-inline'")
                .base_uri("'self'")
                .connect_src("'self'", "data:")
                .frame_src("'self'", 'data:')
                .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(
                status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(
            filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(
            filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append("")
                continue
            
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into Excel format.

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        
        for row in data:
            sheetName = "".join(
                [c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                worksheet = workbook[sheetName]
            except KeyError:
                worksheet = workbook.create_sheet(sheetName)
                rowNums[sheetName] = 1
                # Write headers
                for col_num, header in enumerate(columnNames, 1):
                    worksheet.cell(row=1, column=col_num, value=header)
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=rowNums[sheetName], column=col_num, value=str(cell_value))

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log.

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id)
        except Exception:
            return json.dumps(self.jsonify_error("404", "Scan ID not found")).encode("utf-8")

        if not data:
            return json.dumps(self.jsonify_error("404", "No scan logs found")).encode("utf-8")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([str(x) for x in row])
        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanCorrelations(id)
        except Exception:
            return self.error("Scan ID not found")

        try:
            scan = dbh.scanInstanceGet(id)
        except Exception:
            return self.error("Scan ID not found")

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}-correlations.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(data, headings)

        if filetype.lower() == 'csv':
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}-correlations.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)
            for row in data:
                parser.writerow([str(x) for x in row])
            return fileobj.getvalue()

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format.

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)

        # Get target-level false positives for this scan
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None
        targetFps = set()
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]))
                # Check both per-event FP flag and target-level FPs
                fp_flag = 1 if row[13] or (row[4], row[1]) in targetFps else 0
                rows.append([lastseen, event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])

            fname = "SpiderFoot.xlsx"
            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(
                ["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]))
                # Check both per-event FP flag and target-level FPs
                fp_flag = 1 if row[13] or (row[4], row[1]) in targetFps else 0
                parser.writerow([lastseen, event_type, str(
                    row[3]), str(row[2]), fp_flag, datafield])

            fname = "SpiderFoot.csv"
            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple
        scans.

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        targetFpsPerScan = dict()  # Store target FPs per scan ID
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            # Get target-level false positives for this scan
            target = scaninfo[id][1] if scaninfo[id] else None
            targetFpsPerScan[id] = set()
            if target:
                try:
                    targetFpsPerScan[id] = dbh.targetFalsePositivesForTarget(target)
                except Exception:
                    pass  # Table may not exist in older databases
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]))
                # Check both per-event FP flag and target-level FPs
                scan_id = row[12]
                targetFps = targetFpsPerScan.get(scan_id, set())
                fp_flag = 1 if row[13] or (row[4], row[1]) in targetFps else 0
                rows.append([scaninfo[row[12]][0], lastseen, event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.xlsx"
            else:
                fname = scan_name + "-SpiderFoot.xlsx"

            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type",
                            "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[4]))
                # Check both per-event FP flag and target-level FPs
                scan_id = row[12]
                targetFps = targetFpsPerScan.get(scan_id, set())
                fp_flag = 1 if row[13] or (row[4], row[1]) in targetFps else 0
                parser.writerow([scaninfo[row[12]][0], lastseen, event_type, str(row[3]),
                                str(row[2]), fp_flag, datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.csv"
            else:
                fname = scan_name + "-SpiderFoot.csv"

            cherrypy.response.headers[
                'Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        # Get target-level false positives for this scan
        dbh = SpiderFootDb(self.config)
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None
        targetFps = set()
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[10]))
                # Check both per-event FP flag and target-level FPs
                fp_flag = 1 if row[11] or (row[10], row[1]) in targetFps else 0
                rows.append([row[0], event_type, str(row[3]),
                            str(row[2]), fp_flag, datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(
                ["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                event_type = translate_event_type(str(row[10]))
                # Check both per-event FP flag and target-level FPs
                fp_flag = 1 if row[11] or (row[10], row[1]) in targetFps else 0
                parser.writerow([row[0], event_type, str(
                    row[3]), str(row[2]), fp_flag, datafield])

            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace(
                    "<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.json"
        else:
            fname = scan_name + "-SpiderFoot.json"

        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data or JSON
        """
        # For JSON requests, always return valid JSON (as bytes for CherryPy)
        if gexf == "0":
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            try:
                if not id:
                    return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

                dbh = SpiderFootDb(self.config)
                data = dbh.scanResultEvent(id, filterFp=True)
                scan = dbh.scanInstanceGet(id)

                if not scan:
                    return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

                root = scan[1]
                return SpiderFootHelpers.buildGraphJson([root], data).encode('utf-8')
            except Exception as e:
                self.log.error(f"scanviz JSON error: {e}")
                return json.dumps({'nodes': [], 'edges': []}).encode('utf-8')

        # For GEXF requests
        try:
            if not id:
                return ""

            dbh = SpiderFootDb(self.config)
            data = dbh.scanResultEvent(id, filterFp=True)
            scan = dbh.scanInstanceGet(id)

            if not scan:
                return ""

            scan_name = scan[0]
            root = scan[1]
            fname = f"{scan_name}SpiderFoot.gexf" if scan_name else "SpiderFoot.gexf"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/gexf"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)
        except Exception as e:
            self.log.error(f"scanviz GEXF error: {e}")
            return ""

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "-SpiderFoot.gexf"

        cherrypy.response.headers[
            'Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(
                        key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)
        
        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]
        
        # Validate that we have a valid target
        if not scantarget:
            return self.error(f"Scan {id} has no target defined.")

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = SpiderFootHelpers.targetTypeFromString(
                f'"{scantarget}"')

        # Final validation - ensure we have a valid target type
        if not targetType:
            self.log.error(f"Cannot determine target type for scan rerun: '{scantarget}'")
            return self.error(f"Cannot determine target type for scan rerun. Target '{scantarget}' is not recognized as a valid SpiderFoot target.")

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(
                self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(
            f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            
            # Validate that we have a valid target
            if not scantarget:
                return self.error(f"Scan {id} has no target defined.")
            
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")
                
            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                self.log.error(f"Invalid target type for scan {id}: '{scantarget}' could not be recognized")
                return self.error(f"Invalid target type for scan {id}. Could not recognize '{scantarget}' as a target SpiderFoot supports.")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(
                    self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(
                    f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(
            filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        templ = Template(
            filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)
        
        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        
        # Validate that we have a valid target
        if not scantarget:
            return self.error(f"Scan {id} has no target defined.")
        
        targetType = None
        
        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"
            # Re-check target type after wrapping
            targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                self.log.error(f"Invalid target type for scan {id}: '{scantarget}' could not be recognized")
                return self.error(f"Invalid target type for scan {id}. Could not recognize '{scantarget}' as a target SpiderFoot supports.")

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(
            filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl',
                         lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def workspaces(self: 'SpiderFootWebUi') -> str:
        """Show workspace management page.

        Returns:
            str: Workspace management page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/workspaces.tmpl', lookup=self.lookup)
        return templ.render(pageid='WORKSPACES', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." +
                            mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                value = useropts[opt]
                if not isinstance(value, str):
                    value = str(value)
                cleaned = self.cleanUserInput([value])
                cleanopts[opt] = cleaned[0] if cleaned and len(cleaned) > 0 else ""

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            import logging
            logging.exception("Error processing user input in savesettings")
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    def resultsetfppersist(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str, persist: str = "0") -> str:
        """Set results as false positive with optional target-level persistence.

        This extends resultsetfp to optionally persist false positives at the target level,
        so they will be recognized in future scans of the same target.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1
            persist (str): 0 or 1 - whether to persist at target level

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        target = status[1]  # seed_target

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)

        # Handle target-level persistence
        if ret and persist == "1":
            # Get the event details for each ID to persist at target level
            events = dbh.scanResultEvent(id)
            eventMap = {row[8]: row for row in events}  # hash -> event data

            for resultId in allIds:
                if resultId in eventMap:
                    eventData = eventMap[resultId]
                    eventType = eventData[4]  # type
                    data = eventData[1]  # data

                    if fp == "1":
                        dbh.targetFalsePositiveAdd(target, eventType, data)
                    else:
                        dbh.targetFalsePositiveRemove(target, eventType, data)

        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def targetfplist(self: 'SpiderFootWebUi', target: str = None) -> list:
        """List target-level false positives.

        Args:
            target (str): optional target to filter by

        Returns:
            list: list of target-level false positives
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        fps = dbh.targetFalsePositiveList(target)

        ret = []
        for fp in fps:
            ret.append({
                'id': fp[0],
                'target': fp[1],
                'event_type': fp[2],
                'event_data': fp[3],
                'date_added': fp[4],
                'notes': fp[5]
            })

        return ret

    @cherrypy.expose
    def targetfpadd(self: 'SpiderFootWebUi', target: str, event_type: str, event_data: str, notes: str = None) -> str:
        """Add a target-level false positive.

        Args:
            target (str): target value
            event_type (str): event type
            event_data (str): event data
            notes (str): optional notes

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if not target or not event_type or not event_data:
            return json.dumps(["ERROR", "Missing required parameters."]).encode('utf-8')

        try:
            ret = dbh.targetFalsePositiveAdd(target, event_type, event_data, notes)
            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    def targetfpremove(self: 'SpiderFootWebUi', id: str = None, target: str = None, event_type: str = None, event_data: str = None) -> str:
        """Remove a target-level false positive.

        Can be removed by ID or by target/event_type/event_data combination.

        Args:
            id (str): false positive entry ID
            target (str): target value
            event_type (str): event type
            event_data (str): event data

        Returns:
            str: JSON status
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        try:
            if id:
                ret = dbh.targetFalsePositiveRemoveById(int(id))
            elif target and event_type and event_data:
                ret = dbh.targetFalsePositiveRemove(target, event_type, event_data)
            else:
                return json.dumps(["ERROR", "Must provide either ID or target/event_type/event_data."]).encode('utf-8')

            if ret:
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", str(e)]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        """List all available modules.

        Returns:
            list: list of available modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        
        modlist = list()
        for mod in self.config['__modules__']:
            if "__" in mod:
                continue
            
            modlist.append({
                'name': mod,
                'descr': self.config['__modules__'][mod].get('descr', ''),
                'provides': self.config['__modules__'][mod].get('provides', []),
                'consumes': self.config['__modules__'][mod].get('consumes', []),
                'group': self.config['__modules__'][mod].get('group', [])
            })
            
        return sorted(modlist, key=lambda x: x['name'])

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all available correlation rules.

        Returns:
            list: list of available correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        
        rules = list()
        for rule in self.config.get('__correlationrules__', []):
            rules.append({
                'id': rule.get('id', ''),
                'name': rule.get('name', ''),
                'risk': rule.get('risk', 'UNKNOWN'),
                'description': rule.get('description', '')
            })
            
        return sorted(rules, key=lambda x: x['name'])

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """Ping endpoint for health checks.

        Returns:
            list: status response
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'SpiderFootWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        # Swap the globalscantable for the database handler
        dbh = SpiderFootDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = sf.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = sf.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or ('group' in self.config['__modules__'][mod] and
                                        usecase in self.config['__modules__'][mod]['group']):
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()        # Delete the stdout module in case it crept in
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(
                self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}", exc_info=True)
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        """Vacuum the database."""
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2],
                           html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime(
                    "%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created,
                           started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of
        each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen,
                           row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list or error message
        """
        retdata = []
        dbh = SpiderFootDb(self.config)

        try:
            self.log.debug(f"Fetching correlations for scan {id}")
            corrdata = dbh.scanCorrelationList(id)
            self.log.debug(f"Found {len(corrdata)} correlations")

            if not corrdata:
                self.log.debug(f"No correlations found for scan {id}")
                return retdata

            for row in corrdata:
                # Check if we have a valid row of data
                if len(row) < 6:  # Need at least 6 elements to extract all required fields
                    self.log.error(
                        f"Correlation data format error: missing required fields, got {len(row)} fields")
                    continue

                # Extract specific fields based on their indices
                correlation_id = row[0]
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_id = row[4]
                rule_description = row[5]
                events = row[6] if len(row) > 6 else ""
                created = row[7] if len(row) > 7 else ""

                retdata.append([correlation_id, correlation, rule_name, rule_risk,
                               rule_id, rule_description, events, created])

        except Exception as e:
            self.log.error(
                f"Error fetching correlations for scan {id}: {e}", exc_info=True)
            # Return empty list on error

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results with target-level FP status
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(
                id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        # Get the target for this scan to check target-level FPs
        scanInfo = dbh.scanInstanceGet(id)
        target = scanInfo[1] if scanInfo else None

        # Get all target-level false positives for fast lookup
        targetFps = set()
        if target:
            try:
                targetFps = dbh.targetFalsePositivesForTarget(target)
            except Exception:
                pass  # Table may not exist in older databases

        for row in data:
            lastseen = time.strftime(
                "%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            eventDataRaw = row[1]
            eventTypeRaw = row[4]

            # Check if this result matches a target-level false positive
            isTargetFp = 1 if (eventTypeRaw, eventDataRaw) in targetFps else 0

            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4],
                isTargetFp  # Index 11: target-level false positive flag
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata

    @cherrypy.expose
    def active_maintenance_status(self: 'SpiderFootWebUi') -> str:
        """Display the active maintenance status of the project.

        Returns:
            str: Active maintenance status page HTML
        """
        templ = Template(
            filename='spiderfoot/templates/active_maintenance_status.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def footer(self: 'SpiderFootWebUi') -> str:
        """Display the footer with active maintenance status.

        Returns:
            str: Footer HTML
        """
        templ = Template(
            filename='spiderfoot/templates/footer.tmpl', lookup=self.lookup)
        return templ.render(docroot=self.docroot, version=__version__)

    # Workspace Management API Endpoints
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacelist(self: 'SpiderFootWebUi') -> list:
        """List all workspaces.

        Returns:
            list: List of workspace information
        """
        try:
            workspaces = SpiderFootWorkspace.list_workspaces(self.config)
            return workspaces
        except Exception as e:
            self.log.error(f"Failed to list workspaces: {e}")
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacecreate(self: 'SpiderFootWebUi', name: str, description: str = '') -> dict:
        """Create a new workspace.

        Args:
            name (str): workspace name
            description (str): workspace description

        Returns:
            dict: workspace creation result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, name=name)
            workspace.description = description
            workspace.save_workspace()
            
            return {
                'success': True,
                'workspace_id': workspace.workspace_id,
                'name': workspace.name,
                'description': workspace.description,
                'created_time': workspace.created_time
            }
        except Exception as e:
            self.log.error(f"Failed to create workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceget(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get workspace details.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: workspace information
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            return {
                'success': True,
                'workspace_id': workspace.workspace_id,
                'name': workspace.name,
                'description': workspace.description,
                               'created_time': workspace.created_time,
                'modified_time': workspace.modified_time,
                'targets': workspace.targets,
                'scans': workspace.scans,
                'metadata': workspace.metadata
            }
        except Exception as e:
            self.log.error(f"Failed to get workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceupdate(self: 'SpiderFootWebUi', workspace_id: str, name: str = None, description: str = None) -> dict:
        """Update workspace details.

        Args:
            workspace_id (str): workspace ID
            name (str): new workspace name
            description (str): new workspace description

        Returns:
            dict: update result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            if name is not None:
                workspace.name = name
            if description is not None:
                workspace.description = description
                
            workspace.save_workspace()
            
            return {'success': True, 'message': 'Workspace updated successfully'}
        except Exception as e:
            self.log.error(f"Failed to update workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacedelete(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Delete a workspace.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: deletion result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            workspace.delete_workspace()
            
            return {'success': True, 'message': 'Workspace deleted successfully'}
        except Exception as e:
            self.log.error(f"Failed to delete workspace: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacesummary(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get workspace summary.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: workspace summary
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            summary = workspace.get_workspace_summary()
            
            return {'success': True, 'summary': summary}
        except Exception as e:
            self.log.error(f"Failed to get workspace summary: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceaddtarget(self: 'SpiderFootWebUi', workspace_id: str, target: str, target_type: str = None) -> dict:
        """Add target to workspace.

        Args:
            workspace_id (str): workspace ID
            target (str): target value
            target_type (str): target type

        Returns:
            dict: add target result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            target_id = workspace.add_target(target, target_type)
            
            return {'success': True, 'target_id': target_id, 'message': 'Target added successfully'}
        except Exception as e:
            self.log.error(f"Failed to add target: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceremovetarget(self: 'SpiderFootWebUi', workspace_id: str, target_id: str) -> dict:
        """Remove target from workspace.

        Args:
            workspace_id (str): workspace ID
            target_id (str): target ID

        Returns:
            dict: remove target result
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            success = workspace.remove_target(target_id)
            
            if success:
                return {'success': True, 'message': 'Target removed successfully'}
            else:
                return {'success': False, 'error': 'Target not found'}
        except Exception as e:
            self.log.error(f"Failed to remove target: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspaceimportscans(self: 'SpiderFootWebUi', workspace_id: str, scan_ids: str) -> dict:
        """Import scans into workspace.

        Args:
            workspace_id (str): workspace ID
            scan_ids (str): comma-separated scan IDs

        Returns:
            dict: import result
        """
        try:
            self.log.info(f"[IMPORT] Starting scan import for workspace: {workspace_id}")
            self.log.debug(f"[IMPORT] Raw scan IDs input: {scan_ids}")
            
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            self.log.info(f"[IMPORT] Loaded workspace: {workspace.name}")
            
            # Clean and split scan IDs (handle both comma-separated and line-separated)
            scan_ids_cleaned = scan_ids.replace('\n', ',').replace('\r', '')
            scan_id_list = [sid.strip() for sid in scan_ids_cleaned.split(',') if sid.strip()]
            
            self.log.info(f"[IMPORT] Processed {len(scan_id_list)} scan IDs: {scan_id_list}")
            
            if not scan_id_list:
                return {'success': False, 'error': 'No valid scan IDs provided'}
            
            # Verify scans exist before importing
            dbh = SpiderFootDb(self.config)
            valid_scans = []
            invalid_scans = []
            
            for scan_id in scan_id_list:
                scan_info = dbh.scanInstanceGet(scan_id)
                if scan_info:
                    valid_scans.append(scan_id)
                    self.log.debug(f"[IMPORT] Verified scan {scan_id}: {scan_info[0]}")
                else:
                    invalid_scans.append(scan_id)
                    self.log.warning(f"[IMPORT] Scan {scan_id} not found in database")
            
            if invalid_scans:
                self.log.warning(f"[IMPORT] Invalid scan IDs: {invalid_scans}")
            
            if not valid_scans:
                return {'success': False, 'error': f'No valid scans found. Invalid IDs: {invalid_scans}'}
            
            # Import valid scans
            if len(valid_scans) == 1:
                success = workspace.import_single_scan(valid_scans[0])
                if success:
                    self.log.info(f"[IMPORT] Successfully imported scan {valid_scans[0]}")
                    return {'success': True, 'message': 'Scan imported successfully'}
                else:
                    self.log.error(f"[IMPORT] Failed to import scan {valid_scans[0]}")
                    return {'success': False, 'error': 'Failed to import scan'}
            else:
                results = workspace.bulk_import_scans(valid_scans)
                successful_imports = sum(1 for success in results.values() if success)
                
                self.log.info(f"[IMPORT] Bulk import completed: {successful_imports}/{len(valid_scans)} successful")
                
                message = f'Import completed: {successful_imports} of {len(valid_scans)} scans imported'
                if invalid_scans:
                    message += f'. Invalid scan IDs: {invalid_scans}'
                
                return {
                    'success': True, 
                    'results': results,
                    'message': message,
                    'successful_imports': successful_imports,
                    'total_attempts': len(scan_id_list),
                    'invalid_scans': invalid_scans
                }
        except Exception as e:
            self.log.error(f"[IMPORT] Failed to import scans: {e}")
            import traceback
            self.log.error(f"[IMPORT] Traceback: {traceback.format_exc()}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacemultiscan(self: 'SpiderFootWebUi', workspace_id: str, targets: str, modules: str, scan_name_prefix: str, enable_correlation: str = 'false') -> dict:
        """Start multi-target scan from workspace.

        Args:
            workspace_id (str): workspace ID
            targets (str): JSON string of selected targets
            modules (str): JSON string of selected modules
            scan_name_prefix (str): prefix for scan names
            enable_correlation (str): whether to enable correlation

        Returns:
            dict: multi-target scan result
        """
        self.log.info(f"[MULTISCAN] Starting multi-target scan for workspace: {workspace_id}")
        self.log.debug(f"[MULTISCAN] Input parameters - targets: {targets}, modules: {modules}, prefix: {scan_name_prefix}")
        
        try:
            self.log.debug(f"[MULTISCAN] Importing startSpiderFootScanner...")
            from sfscan import startSpiderFootScanner
            self.log.debug(f"[MULTISCAN] Import successful")
            
            # Try to load existing workspace, or create a new one if it doesn't exist
            self.log.debug(f"[MULTISCAN] Attempting to load workspace: {workspace_id}")
            try:
                workspace = SpiderFootWorkspace(self.config, workspace_id)
                self.log.info(f"[MULTISCAN] Successfully loaded existing workspace: {workspace_id}")
            except (ValueError, Exception) as e:
                # Workspace doesn't exist, create a new one
                self.log.info(f"[MULTISCAN] Workspace {workspace_id} not found ({e}), creating new one")
                try:
                    workspace = SpiderFootWorkspace(self.config, name=f"Workspace_{workspace_id}")
                    workspace.workspace_id = workspace_id  # Override the generated ID                    workspace.save_workspace()
                    self.log.info(f"[MULTISCAN] Successfully created new workspace: {workspace_id}")
                except Exception as create_error:
                    self.log.error(f"[MULTISCAN] Failed to create workspace: {create_error}")
                    raise
            
            # Parse targets and modules
            self.log.debug("[MULTISCAN] Parsing JSON input data...")
            try:
                target_list = json.loads(targets)
                self.log.debug(f"[MULTISCAN] Parsed {len(target_list)} targets: {[t.get('value', 'unknown') for t in target_list]}")
            except Exception as e:
                self.log.error(f"[MULTISCAN] Failed to parse targets JSON: {e}")
                raise ValueError(f"Invalid targets JSON: {e}")
            
            try:
                module_list = json.loads(modules)
                self.log.debug(f"[MULTISCAN] Parsed {len(module_list)} modules: {module_list}")
            except Exception as e:
                self.log.error(f"[MULTISCAN] Failed to parse modules JSON: {e}")
                raise ValueError(f"Invalid modules JSON: {e}")
            
            scan_ids = []
            
            self.log.info(f"[MULTISCAN] Starting scan loop for {len(target_list)} targets")
            
            # Start a scan for each target
            for i, target in enumerate(target_list):
                self.log.debug(f"[MULTISCAN] Processing target {i+1}/{len(target_list)}: {target}")
                
                target_value = target['value']
                target_type = target.get('type', '')
                
                self.log.debug(f"[MULTISCAN] Target value: {target_value}, type: {target_type}")
                
                # If target type is not provided or empty, detect it
                if not target_type:
                    self.log.debug(f"[MULTISCAN] Detecting target type for: {target_value}")
                    target_type = SpiderFootHelpers.targetTypeFromString(target_value)
                    if target_type is None:
                        self.log.error(f"[MULTISCAN] Could not determine target type for {target_value}")
                        continue
                    else:
                        self.log.debug(f"[MULTISCAN] Detected target type: {target_type}")
                
                # Normalize target value like other scan methods
                original_value = target_value
                if target_type in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
                    target_value = target_value.replace("\"", "")
                else:
                    target_value = target_value.lower()
                
                if original_value != target_value:
                    self.log.debug(f"[MULTISCAN] Normalized target value: {original_value} -> {target_value}")
                
                # Generate scan name
                scan_name = f"{scan_name_prefix} - {target_value}"
                self.log.debug(f"[MULTISCAN] Generated scan name: {scan_name}")
                
                # Create module configuration list (like in working examples)
                modlist = module_list.copy()
                self.log.debug(f"[MULTISCAN] Initial module list: {modlist}")
                
                # Add our mandatory storage module
                if "sfp__stor_db" not in modlist:
                    modlist.append("sfp__stor_db")
                    self.log.debug("[MULTISCAN] Added mandatory sfp__stor_db module")
                
                # Delete the stdout module in case it crept in
                if "sfp__stor_stdout" in modlist:
                    modlist.remove("sfp__stor_stdout")
                    self.log.debug("[MULTISCAN] Removed sfp__stor_stdout module")
                
                self.log.debug(f"[MULTISCAN] Final module list: {modlist}")
                
                # Create configuration copy for this scan
                self.log.debug("[MULTISCAN] Creating configuration copy...")
                cfg = deepcopy(self.config)
                
                # Start the scan using the correct signature
                scanId = SpiderFootHelpers.genScanInstanceId()
                self.log.info(f"[MULTISCAN] Generated scan ID {scanId} for target {target_value}")                
                try:
                    self.log.debug(f"[MULTISCAN] Starting process for scan {scanId}")
                    # Use multiprocessing like the working examples
                    # startSpiderFootScanner signature: (loggingQueue, *args)
                    # where args are: (scanName, scanId, targetValue, targetType, moduleList, globalOpts)
                    p = mp.Process(target=startSpiderFootScanner, args=(
                        self.loggingQueue, scan_name, scanId, target_value, target_type, modlist, cfg))
                    p.daemon = True
                    p.start()
                    self.log.info(f"[MULTISCAN] Successfully started process for scan {scanId}")
                    
                    scan_ids.append(scanId)
                    
                    # Wait a moment for the scan to initialize in the database
                    import time
                    time.sleep(0.5)
                    
                    # Import the scan into the workspace
                    self.log.debug(f"[MULTISCAN] Importing scan {scanId} into workspace {workspace_id}")
                    workspace.import_single_scan(scanId, {
                        'source': 'multi_target_scan',
                        'scan_name_prefix': scan_name_prefix,
                        'target_id': target.get('target_id', 'unknown'),
                        'imported_time': time.time()
                    })
                    self.log.debug(f"[MULTISCAN] Successfully imported scan {scanId} into workspace")
                    
                except Exception as e:
                    self.log.error(f"[MULTISCAN] Failed to start scan for target {target_value}: {e}")
                    import traceback
                    self.log.error(f"[MULTISCAN] Traceback: {traceback.format_exc()}")
                    continue
            
            self.log.info(f"[MULTISCAN] Scan loop completed. Started {len(scan_ids)} out of {len(target_list)} scans")
            
            if scan_ids:
                message = f"Started {len(scan_ids)} scans successfully"
                if enable_correlation.lower() == 'true':
                    message += ". Correlation analysis will be available once scans complete"
                
                self.log.info(f"[MULTISCAN] Success: {message}")
                return {
                    'success': True,
                    'message': message,
                    'scan_ids': scan_ids,
                    'workspace_id': workspace_id
                }
            else:
                error_msg = 'Failed to start any scans'
                self.log.error(f"[MULTISCAN] {error_msg}")
                return {'success': False, 'error': error_msg}                
        except Exception as e:
            self.log.error(f"[MULTISCAN] Failed to start multi-target scan: {e}")
            import traceback
            self.log.error(f"[MULTISCAN] Traceback: {traceback.format_exc()}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacemcpreport(self: 'SpiderFootWebUi', workspace_id: str, report_type: str, format: str = 'json', 
                          include_correlations: str = 'true', include_threat_intel: str = 'true', 
                          include_recommendations: str = 'true', tlp_level: str = 'amber') -> dict:
        """Generate MCP CTI report for workspace.

        Args:
            workspace_id (str): workspace ID
            report_type (str): type of report (threat_assessment, ioc_analysis, etc.)
            format (str): output format (json, markdown, pdf, html)
            include_correlations (str): include correlation analysis
            include_threat_intel (str): include threat intelligence context
            include_recommendations (str): include security recommendations
            tlp_level (str): Traffic Light Protocol level

        Returns:
            dict: {'success': bool, 'download_url': str, 'error': str}        """
        try:
            # Validate workspace exists
            workspace = SpiderFootWorkspace(self.config, workspace_id)

            # Get workspace scans for report data
            if not workspace.scans:
                return {'success': True, 'correlations': [], 'message': 'Need at least 2 scans for cross-correlation analysis'}
            
            # Import MCP integration
            try:
                from spiderfoot.mcp_integration import SpiderFootMCPClient
                mcp_client = SpiderFootMCPClient(self.config)
            except ImportError:
                return {'success': False, 'error': 'MCP integration not available'}

            # Prepare report configuration
            report_config = {
                'workspace_id': workspace_id,
                'workspace_name': workspace.get('name', 'Unnamed'),
                'report_type': report_type,
                'format': format,
                'options': {
                    'include_correlations': include_correlations.lower() == 'true',
                    'include_threat_intel': include_threat_intel.lower() == 'true',
                    'include_recommendations': include_recommendations.lower() == 'true',
                    'tlp_level': tlp_level
                },
                'scan_ids': [scan['scan_id'] for scan in workspace.scans]
            }

            # Generate report asynchronously (this is a placeholder for actual MCP integration)
            # In a real implementation, this would call the MCP server
            import uuid
            import time
            report_id = str(uuid.uuid4())
            timestamp = int(time.time())
            
            # Create download URL (placeholder - would be actual file in production)
            download_url = f"/workspacereportdownload?report_id={report_id}&workspace_id={workspace_id}&format={format}"
            
            self.log.info(f"Generated MCP report for workspace {workspace_id}: {report_id}")
            
            return {
                'success': True,
                'report_id': report_id,
                'download_url': download_url,
                'message': f'MCP {report_type} report generated successfully'
            }

        except Exception as e:
            self.log.error(f"Failed to generate MCP report: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out() 
    def workspacetiming(self: 'SpiderFootWebUi', workspace_id: str, timezone: str = None, 
                       default_start_time: str = None, retention_period: str = None,
                       auto_scheduling: str = None, business_hours_only: str = None,
                       enable_throttling: str = None, business_start: str = None, 
                       business_end: str = None) -> dict:
        """Get or set workspace timing configuration.

        Args:
            workspace_id (str): workspace ID
            timezone (str): workspace timezone
            default_start_time (str): default scan start time (HH:MM)
            retention_period (str): data retention period in days
            auto_scheduling (str): enable automatic scheduling
            business_hours_only (str): restrict scans to business hours
            enable_throttling (str): enable scan rate throttling
            business_start (str): business hours start time (HH:MM)
            business_end (str): business hours end time (HH:MM)

        Returns:
            dict: timing configuration or success status
        """
        try:
            # Validate workspace exists
            workspace = SpiderFootWorkspace(self.config, workspace_id)

            # If this is a GET request (no parameters provided for setting)
            if timezone is None and default_start_time is None:
                # Return current timing configuration
                timing_config = workspace.metadata.get('timing_config', {})
                return {
                    'success': True,
                    'timezone': timing_config.get('timezone', 'UTC'),
                    'default_start_time': timing_config.get('default_start_time', '09:00'),
                    'retention_period': timing_config.get('retention_period', '90'),
                    'auto_scheduling': timing_config.get('auto_scheduling', False),
                    'business_hours_only': timing_config.get('business_hours_only', False),
                    'enable_throttling': timing_config.get('enable_throttling', True),
                    'business_start': timing_config.get('business_start', '08:00'),
                    'business_end': timing_config.get('business_end', '18:00')
                }

            # This is a POST request - update timing configuration
            timing_config = {
                'timezone': timezone or 'UTC',
                'default_start_time': default_start_time or '09:00', 
                'retention_period': int(retention_period) if retention_period else 90,
                'auto_scheduling': auto_scheduling == 'true' if auto_scheduling else False,
                'business_hours_only': business_hours_only == 'true' if business_hours_only else False,
                'enable_throttling': enable_throttling != 'false',  # Default to True
                'business_start': business_start or '08:00',
                'business_end': business_end or '18:00',
                'updated_time': time.time()
            }            # Update workspace with timing configuration
            import time
            workspace.metadata['timing_config'] = timing_config
            workspace.save_workspace()

            self.log.info(f"Updated timing configuration for workspace {workspace_id}")
            return {'success': True, 'message': 'Timing configuration updated successfully'}

        except Exception as e:
            self.log.error(f"Failed to handle workspace timing: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    def workspacereportdownload(self: 'SpiderFootWebUi', report_id: str, workspace_id: str, format: str = 'json'):
        """Download generated MCP report.

        Args:
            report_id (str): report identifier
            workspace_id (str): workspace ID
            format (str): report format

        Returns:
            File download or error page
        """
        try:
            # Validate workspace access
            workspace = SpiderFootWorkspace(self.config, workspace_id)            # Generate sample report content (placeholder)
            import json
            from datetime import datetime
            
            sample_report = {
                'report_id': report_id,
                'workspace_id': workspace_id,
                'workspace_name': workspace.name,
                'generated_time': datetime.now().isoformat(),
                'report_type': 'MCP CTI Report',
                'format': format,
                'status': 'This is a placeholder MCP report. Integration with actual MCP server required.',
                'summary': {
                    'total_targets': len(workspace.targets),
                    'total_scans': len(workspace.scans),
                    'risk_level': 'Medium',
                    'key_findings': [
                        'Placeholder finding 1',
                        'Placeholder finding 2', 
                        'Placeholder finding 3'
                    ]
                }
            }            # Set appropriate headers for download
            cherrypy.response.headers['Content-Type'] = 'application/octet-stream'
            cherrypy.response.headers['Content-Disposition'] = f'attachment; filename="mcp_report_{report_id}.{format}"'

            if format == 'json':
                return json.dumps(sample_report, indent=2)
            elif format == 'markdown':
                md_content = f"""# MCP CTI Report
                
**Report ID:** {report_id}
**Workspace:** {workspace.name}
**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Summary
- Total Targets: {len(workspace.targets)}
- Total Scans: {len(workspace.scans)}
- Risk Level: Medium

## Status
This is a placeholder MCP report. Integration with actual MCP server required.
"""
                return md_content
            else:
                return json.dumps(sample_report, indent=2)
                
        except Exception as e:
            self.log.error(f"Failed to download report: {e}")
    @cherrypy.expose
    def documentation(self: 'SpiderFootWebUi', doc: str = None, q: str = None) -> str:
        """
        Render documentation from the documentation/ folder as HTML, including subfolders.
        """
        self.log.debug("Documentation endpoint called with doc=%s, q=%s", doc, q)
        import re
        doc_dir = os.path.join(os.path.dirname(__file__), 'documentation')
        doc_dir = os.path.abspath(doc_dir)
        doc_index = []
        selected_file = None
        content = ''
        search_results = []
        search_query = q or ''
        toc_html = ''
        breadcrumbs = []
        last_updated = ''
        author = ''
        version_dirs = []
        current_version = 'latest'
        related = []
        try:
            # Recursively find all .md files
            md_files = []
            for root, dirs, files in os.walk(doc_dir):
                for fname in files:
                    if fname.endswith('.md'):
                        rel_path = os.path.relpath(os.path.join(root, fname), doc_dir)
                        rel_path = rel_path.replace('\\', '/')  # For Windows compatibility
                        md_files.append(rel_path)
            # Use README.md table for sidebar if present
            readme_path = os.path.join(doc_dir, 'README.md')
            sidebar_entries = []
            if os.path.exists(readme_path):
                with open(readme_path, encoding='utf-8') as f:
                    readme_content = f.read()
                # Extract table rows: | Section | File | Icon |
                table_rows = re.findall(r'\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|', readme_content)
                for section, file, icon in table_rows:
                    # Skip header row
                    if section.lower() == 'section':
                        continue
                    # Only add if file exists in md_files
                    if file.strip() in md_files:
                        sidebar_entries.append({
                            'file': file.strip(),
                            'title': section.strip(),
                            'icon': f'fa {icon.strip()}'
                        })
            # Fallback: use all .md files
            if not sidebar_entries:
                for rel_path in sorted(md_files):
                    title = rel_path.replace('.md', '').replace('_', ' ').replace('/', ' / ').title()
                    sidebar_entries.append({
                        'file': rel_path,
                        'title': title,
                        'icon': 'fa fa-file-text-o'
                    })
            doc_index = sidebar_entries
            # Determine which file to show
            if doc and doc.endswith('.md') and doc in md_files:
                selected_file = doc
            elif doc_index:
                selected_file = doc_index[0]['file']
            # Read and render the selected file
            if selected_file:
                file_path = os.path.join(doc_dir, selected_file)
                try:
                    with open(file_path, encoding='utf-8') as f:
                        raw_content = f.read()
                    content = markdown.markdown(
                        raw_content,
                        extensions=['extra', 'toc', 'tables', 'fenced_code']
                    )
                    # --- POST-PROCESS: Rewrite .md links to /documentation?doc=... ---
                    def md_link_rewrite(match):
                        text, url = match.group(1), match.group(2)
                        if url.endswith('.md'):
                            # Remove leading ./ or / if present
                            url = url.lstrip('./')
                            return f'<a href="/documentation?doc={url}">{text}</a>'
                        return match.group(0)
                    content = re.sub(r'<a href=["\']([^"\']+\.md)["\']>(.*?)</a>',
                                     lambda m: f'<a href="/documentation?doc={m.group(1)}">{m.group(2)}</a>',
                                     content)
                    # Also handle Markdown links rendered as <a href="modules/sfp_virustotal.md">...</a>
                    content = re.sub(r'<a href=["\'](modules/[^"\']+\.md)["\']>(.*?)</a>',
                                     lambda m: f'<a href="/documentation?doc={m.group(1)}">{m.group(2)}</a>',
                                     content)
                except Exception as e:
                    self.log.error("Failed to load documentation file %s: %s", file_path, e)
                    content = (
                        '<div class="alert alert-danger">'
                        f'Failed to load documentation: {e}'
                        '</div>'
                    )
            # Search functionality
            if search_query:
                for entry in doc_index:
                    file_path = os.path.join(doc_dir, entry['file'])
                    try:
                        with open(file_path, encoding='utf-8') as f:
                            text = f.read()
                        if (
                            search_query.lower() in text.lower()
                            or search_query.lower() in entry['title'].lower()
                        ):
                            search_results.append(entry)
                    except Exception as e:
                        self.log.warning("Error searching documentation file %s: %s", file_path, e)
                        continue
            # Breadcrumbs (simple: Home > Current)
            breadcrumbs = [
                {
                    'url': self.docroot + '/documentation',
                    'title': 'Documentation'
                }
            ]
            if selected_file:
                breadcrumbs.append({
                    'url': (
                        self.docroot
                        + '/documentation?doc='
                        + selected_file
                    ),
                    'title': (
                        selected_file.replace('.md', '')
                        .replace('_', ' ')
                        .replace('/', ' / ')
                        .title()
                    )
                })
            # Render template
            templ = Template(
                filename='spiderfoot/templates/documentation.tmpl',
                lookup=self.lookup
            )
            # Provide a dummy highlight function if not searching
            def highlight(text, query):
                import re
                if not text or not query:
                    return text
                pattern = re.compile(re.escape(query), re.IGNORECASE)
                return pattern.sub(lambda m: f'<mark>{m.group(0)}</mark>', text)
            return templ.render(
                docroot=self.docroot,
                doc_index=doc_index,
                selected_file=selected_file,
                content=content,
                search_query=search_query,
                search_results=search_results,
                toc_html=toc_html,
                breadcrumbs=breadcrumbs,
                last_updated=last_updated,
                author=author,
                version_dirs=version_dirs,
                current_version=current_version,
                related=related,
                version=__version__,
                pageid="DOCUMENTATION",
                highlight=highlight
            )
        except Exception as e:
            self.log.error("Error in documentation endpoint: %s", e, exc_info=True)
            return (
                '<div class="alert alert-danger">'
            f'Error loading documentation: {e}'
            '</div>'
        )
    @cherrypy.expose
    def workspacedetails(self: 'SpiderFootWebUi', workspace_id: str) -> str:
        """Enhanced workspace details page.

        Args:
            workspace_id (str): workspace ID

        Returns:
            str: workspace details page HTML
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            # Refresh workspace to get latest scan data
            workspace.load_workspace()
            
            # Get workspace summary and scan details
            dbh = SpiderFootDb(self.config)
            scan_details = []
            
            for scan in workspace.scans:
                scan_info = dbh.scanInstanceGet(scan['scan_id'])
                if scan_info:
                    scan_details.append({
                        'scan_id': scan['scan_id'],
                        'name': scan_info[0],
                        'target': scan_info[1],
                        'status': scan_info[5],
                        'created': scan_info[2],
                        'started': scan_info[3],
                        'ended': scan_info[4],
                        'imported_time': scan.get('imported_time', 0)
                    })
            
            templ = Template(filename='spiderfoot/templates/workspace_details.tmpl', lookup=self.lookup)
            return templ.render(
                workspace=workspace,
                scan_details=scan_details,
                docroot=self.docroot,
                version=__version__,
                pageid="WORKSPACE_DETAILS"
            )
            
        except Exception as e:
            self.log.error(f"Error loading workspace details: {e}")
            return self.error(f"Error loading workspace details: {e}")
        
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacescancorrelations(self: 'SpiderFootWebUi', workspace_id: str) -> dict:
        """Get cross-scan correlations for a workspace.

        Args:
            workspace_id (str): workspace ID

        Returns:
            dict: correlation analysis results
        """
        try:
            workspace = SpiderFootWorkspace(self.config, workspace_id)
            
            if not workspace.scans or len(workspace.scans) < 2:
                return {'success': True, 'correlations': [], 'message': 'Need at least 2 scans for cross-correlation analysis'}
            
            dbh = SpiderFootDb(self.config)
            correlations = []
              # Get correlations for each scan
            finished_scans = 0
            for scan in workspace.scans:
                # Check if scan is finished before looking for correlations
                scan_info = dbh.scanInstanceGet(scan['scan_id'])
                if scan_info and scan_info[5] == 'FINISHED':
                    finished_scans += 1
                    scan_correlations = dbh.scanCorrelationList(scan['scan_id'])
                    for corr in scan_correlations:
                        correlations.append({
                            'scan_id': scan['scan_id'],
                            'correlation_id': corr[0],
                            'correlation': corr[1],
                            'rule_name': corr[2],
                            'rule_risk': corr[3],
                            'rule_id': corr[4],
                            'rule_description': corr[5],
                            'created': corr[7] if len(corr) > 7 else ''
                        })
            
            # Check if we have enough finished scans for correlation analysis
            if finished_scans < 2:
                return {
                    'success': True, 
                    'correlations': [], 
                    'correlation_groups': {},
                    'total_correlations': 0,
                    'cross_scan_patterns': 0,
                    'finished_scans': finished_scans,
                    'total_scans': len(workspace.scans),
                    'message': f'Need at least 2 finished scans for correlation analysis. Currently have {finished_scans} finished out of {len(workspace.scans)} total scans.'
                }
              # Group correlations by rule type
            correlation_groups = {}
            for corr in correlations:
                rule_name = corr['rule_name']
                if rule_name not in correlation_groups:
                    correlation_groups[rule_name] = []
                correlation_groups[rule_name].append(corr)
            
            return {
                'success': True,
                'correlations': correlations,
                'correlation_groups': correlation_groups,
                'total_correlations': len(correlations),
                'cross_scan_patterns': len(correlation_groups),
                'finished_scans': finished_scans,
                'total_scans': len(workspace.scans)
            }
            
        except Exception as e:
            self.log.error(f"Error getting workspace correlations: {e}")
            return {'success': False, 'error': str(e)}

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def workspacescanresults(self: 'SpiderFootWebUi', workspace_id: str, scan_id: str = None, event_type: str = None, limit: int = 100) -> dict:
        """Get scan results for workspace scans.

        Args:
            workspace_id (str): workspace ID
            scan_id (str): specific scan ID (optional)
            event_type (str): filter by event type (optional)
            limit (int): maximum results to return

        Returns:
            dict: scan results data
        """
        try:
            # Convert limit to integer if it's passed as string from HTTP request
            if isinstance(limit, str):
                try:
                    limit = int(limit)
                except (ValueError, TypeError):
                    limit = 100  # fallback to default

            # Ensure limit is positive and reasonable
            if not isinstance(limit, int) or limit <= 0:
                limit = 100
            elif limit > 10000:  # Cap at reasonable maximum
                limit = 10000

            workspace = SpiderFootWorkspace(self.config, workspace_id)
            dbh = SpiderFootDb(self.config)

            if scan_id:
                # Get results for specific scan
                scan_ids = [scan_id]
            else:
                # Get results for all workspace scans
                scan_ids = [scan['scan_id'] for scan in workspace.scans]

            all_results = []
            scan_summaries = {}

            for sid in scan_ids:
                # Get scan summary
                summary = dbh.scanResultSummary(sid, 'type')
                scan_summaries[sid] = summary

                # Get recent events
                if event_type:
                    events = dbh.scanResultEvent(sid, event_type, False)
                else:
                    events = dbh.scanResultEvent(sid, 'ALL', False)

                # Limit results per scan
                events = events[:limit] if events else []

                for event in events:
                    all_results.append({
                        'scan_id': sid,
                        'timestamp': event[0],
                        'event_type': event[1],
                        'event_data': event[2],
                        'source_module': event[3],
                        'source_event': event[4] if len(event) > 4 else '',
                        'false_positive': event[8] if len(event) > 8 else False
                    })

            return {
                'success': True,
                'results': all_results[:limit],  # Apply overall limit
                'scan_summaries': scan_summaries,
                'total_results': len(all_results),
                'workspace_id': workspace_id
            }
            
        except Exception as e:
            self.log.error(f"Error getting workspace scan results: {e}")
            return {'success': False, 'error': str(e)}