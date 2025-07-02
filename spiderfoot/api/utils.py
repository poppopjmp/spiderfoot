"""
Utility functions for SpiderFoot API (CSV/Excel builders, etc.)
"""
import html
import openpyxl
import string
from io import BytesIO

def clean_user_input(input_list: list) -> list:
    ret = []
    for item in input_list:
        if isinstance(item, str):
            c = html.escape(item, quote=False)
            ret.append(c)
        else:
            ret.append(item)
    return ret

def build_excel(data: list, column_names: list, sheet_name_index: int = 0) -> str:
    row_nums = dict()
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    column_names.pop(sheet_name_index)
    allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
    for row in data:
        if len(row) < len(column_names):
            continue
        sheet_name = row[sheet_name_index]
        sheet_name = ''.join(c for c in sheet_name if c in allowed_sheet_chars)[:31]
        if sheet_name not in workbook.sheetnames:
            if len(workbook.sheetnames) == 1 and workbook.sheetnames[0] == 'Sheet':
                sheet = default_sheet
                sheet.title = sheet_name
            else:
                sheet = workbook.create_sheet(sheet_name)
            row_nums[sheet_name] = 1
            for col_num, column_title in enumerate(column_names, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            row_nums[sheet_name] += 1
        else:
            sheet = workbook[sheet_name]
        for col_num, cell_value in enumerate([v for i, v in enumerate(row) if i != sheet_name_index], 1):
            sheet.cell(row=row_nums[sheet_name], column=col_num, value=cell_value)
        row_nums[sheet_name] += 1
    if row_nums:
        workbook._sheets.sort(key=lambda ws: ws.title)
    with BytesIO() as f:
        workbook.save(f)
        f.seek(0)
        return f.read()
