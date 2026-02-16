"""
Scan router — all endpoints delegate to ``ScanService``.

Cycle 27 migrated 5 core endpoints.  Cycle 29 completes the migration
of all 25 endpoints, eliminating every raw ``SpiderFootDb`` call.
"""

from __future__ import annotations

import csv
import json
import logging
import multiprocessing as mp
import time
from copy import deepcopy
from io import BytesIO, StringIO

import openpyxl
from fastapi import APIRouter, BackgroundTasks, Body, Depends, HTTPException, Query
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel, Field

from spiderfoot import SpiderFootHelpers
from spiderfoot.scan_service.scanner import startSpiderFootScanner
from spiderfoot.scan.scan_service_facade import ScanService, ScanServiceError
from spiderfoot.sflib.core import SpiderFoot

from ..dependencies import get_app_config, get_api_key, optional_auth, get_scan_service
from ..pagination import PaginationParams, paginate
from ..schemas import (
    ScanCreateResponse,
    ScanDeleteResponse,
    ScanStopResponse,
    ScanMetadataResponse,
    ScanNotesResponse,
    ScanRerunResponse,
    ScanCloneResponse,
    ScanTagsResponse,
    MessageResponse,
)

# Scan lifecycle hooks — best-effort, non-blocking
try:
    from spiderfoot.scan.scan_hooks import get_scan_hooks
    _hooks = get_scan_hooks()
except Exception as e:
    _hooks = None  # type: ignore[assignment]

router = APIRouter()
log = logging.getLogger(__name__)

api_key_dep = Depends(get_api_key)
optional_auth_dep = Depends(optional_auth)


# -----------------------------------------------------------------------
# Request models
# -----------------------------------------------------------------------

class ScanRequest(BaseModel):
    """Data model for a scan creation request."""
    name: str = Field(..., description="Name of the scan")
    target: str = Field(..., description="Target for the scan")
    modules: list[str] | None = Field(None, description="List of module names to run")
    type_filter: list[str] | None = Field(None, description="List of event types to include")
    engine: str | None = Field(None, description="Scan engine profile name (from /api/engines)")
    profile: str | None = Field(None, description="Scan profile name (e.g. 'tools-only', 'quick-recon')")


class ScheduleCreateRequest(BaseModel):
    """Request to create a recurring scan schedule."""
    name: str = Field(..., description="Schedule name")
    target: str = Field(..., description="Scan target")
    interval_minutes: int = Field(0, ge=0, description="Run every N minutes (0 = one-shot)")
    run_at: float | None = Field(None, description="Unix timestamp for one-shot execution")
    modules: list[str] | None = Field(None, description="Module list")
    type_filter: list[str] | None = Field(None, description="Event type filter")
    max_runs: int = Field(0, ge=0, description="Max runs (0 = unlimited)")
    description: str = ""
    tags: list[str] | None = None


# -----------------------------------------------------------------------
# Scan Profiles
# -----------------------------------------------------------------------

@router.get("/scan-profiles", summary="List available scan profiles")
def list_scan_profiles():
    """Return all built-in and custom scan profiles."""
    try:
        from spiderfoot.scan.scan_profile import get_profile_manager
        pm = get_profile_manager()
        profiles = []
        for p in pm.list():
            profiles.append({
                "name": p.name,
                "display_name": p.display_name,
                "description": p.description,
                "category": p.category.value if hasattr(p.category, "value") else str(p.category),
                "module_count": len(p.include_modules),
                "modules": sorted(p.include_modules),
                "tags": p.tags,
                "max_threads": p.max_threads,
                "timeout_minutes": p.timeout_minutes,
            })
        return {"profiles": profiles, "total": len(profiles)}
    except Exception as e:
        log.warning("Failed to load scan profiles: %s", e)
        return {"profiles": [], "total": 0}


@router.get("/scan-profiles/{profile_name}", summary="Get a specific scan profile")
def get_scan_profile(profile_name: str):
    """Return details of a specific scan profile."""
    try:
        from spiderfoot.scan.scan_profile import get_profile_manager
        pm = get_profile_manager()
        p = pm.get(profile_name)
        if not p:
            raise HTTPException(status_code=404, detail=f"Profile '{profile_name}' not found")
        return {
            "name": p.name,
            "display_name": p.display_name,
            "description": p.description,
            "category": p.category.value if hasattr(p.category, "value") else str(p.category),
            "module_count": len(p.include_modules),
            "modules": sorted(p.include_modules),
            "include_flags": p.include_flags,
            "exclude_flags": p.exclude_flags,
            "tags": p.tags,
            "max_threads": p.max_threads,
            "timeout_minutes": p.timeout_minutes,
        }
    except HTTPException:
        raise
    except Exception as e:
        log.warning("Failed to load scan profile '%s': %s", profile_name, e)
        raise HTTPException(status_code=500, detail="Failed to load profile")


# -----------------------------------------------------------------------
# Background task helper
# -----------------------------------------------------------------------

def start_scan_background(
    scan_id: str,
    scan_name: str,
    target: str,
    target_type: str,
    modules: list,
    type_filter: list,
    config: dict,
) -> None:
    """Launch a SpiderFoot scan in the background for the given target and modules.

    NOTE: This is intentionally a sync function (not async) so that FastAPI's
    BackgroundTasks runs it in a thread pool instead of blocking the event loop.

    If Celery is available, dispatches to the Celery worker instead of running
    in-process.  Falls back to multiprocessing if Celery is unreachable.
    """
    # Try Celery-based execution first (v5.4.0+)
    try:
        from spiderfoot.celery_app import is_celery_available
        if is_celery_available():
            from spiderfoot.tasks.scan import run_scan
            run_scan.apply_async(
                kwargs={
                    "scan_name": scan_name,
                    "scan_id": scan_id,
                    "target_value": target,
                    "target_type": target_type,
                    "module_list": modules,
                    "global_opts": config,
                },
                task_id=scan_id,
                queue="scan",
            )
            log.info("Scan %s dispatched to Celery worker", scan_id)
            return
    except Exception as e:
        log.warning("Celery dispatch failed for scan %s, falling back to in-process: %s", scan_id, e)

    # Fallback: run in-process via multiprocessing
    import multiprocessing as mp
    from spiderfoot.observability.logger import logListenerSetup
    try:
        logging_queue = mp.Queue()
        logListenerSetup(logging_queue, config)
        startSpiderFootScanner(
            logging_queue, scan_name, scan_id, target, target_type, modules, config,
        )
    except Exception as e:
        log.error("Failed to start scan %s: %s", scan_id, e)


# -----------------------------------------------------------------------
# Static routes (MUST come before parameterized routes)
# -----------------------------------------------------------------------


@router.get("/scans/export-multi")
async def export_scan_json_multi(
    ids: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> StreamingResponse:
    """Export event results for multiple scans as JSON."""
    scaninfo: list = []
    scan_name = ""

    for scan_id in ids.split(","):
        scan_id = scan_id.strip()
        if not scan_id:
            continue
        record = svc.get_scan(scan_id)
        if record is None:
            continue
        scan_name = record.name
        for row in svc.get_events(scan_id):
            if row[4] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
            scaninfo.append({
                "data": event_data,
                "event_type": row[4],
                "module": str(row[3]),
                "source_data": str(row[2]),
                "false_positive": row[13] if len(row) > 13 else None,
                "last_seen": lastseen,
                "scan_name": record.name,
                "scan_target": record.target,
            })

    id_list = [s.strip() for s in ids.split(",") if s.strip()]
    if len(id_list) > 1 or not scan_name:
        fname = "SpiderFoot.json"
    else:
        fname = f"{scan_name}-SpiderFoot.json"

    return StreamingResponse(
        iter([json.dumps(scaninfo, ensure_ascii=False)]),
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}", "Pragma": "no-cache"},
    )


@router.get("/scans/viz-multi")
async def export_scan_viz_multi(
    ids: str,
    gexf: str = "1",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> Response:
    """Export entities from multiple scans in GEXF format."""
    if not ids:
        raise HTTPException(status_code=400, detail="No scan IDs provided")

    data: list = []
    roots: list = []
    scan_name = ""

    for scan_id in ids.split(","):
        scan_id = scan_id.strip()
        if not scan_id:
            continue
        record = svc.get_scan(scan_id)
        if record is None:
            continue
        data += svc.get_events(scan_id, filter_fp=True)
        roots.append(record.target)
        scan_name = record.name

    if not data:
        raise HTTPException(status_code=404, detail="No data found for these scans")

    if gexf == "0":
        raise HTTPException(status_code=501, detail="Graph JSON for multi-scan not implemented")

    id_list = [s.strip() for s in ids.split(",") if s.strip()]
    fname = (
        f"{scan_name}-SpiderFoot.gexf"
        if len(id_list) == 1 and scan_name
        else "SpiderFoot.gexf"
    )
    gexf_data = SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)
    return Response(
        gexf_data,
        media_type="application/gexf",
        headers={"Content-Disposition": f"attachment; filename={fname}", "Pragma": "no-cache"},
    )


@router.post("/scans/rerun-multi")
async def rerun_scan_multi(
    ids: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Rerun multiple scans."""
    config = get_app_config()
    cfg = deepcopy(config.get_config())
    dbh = svc.dbh
    new_scan_ids: list = []

    for scan_id in ids.split(","):
        scan_id = scan_id.strip()
        if not scan_id:
            continue
        record = svc.get_scan(scan_id)
        if record is None:
            continue
        scanconfig = dbh.scanConfigGet(scan_id)
        if not record.target or not scanconfig:
            continue
        modlist = scanconfig["_modulesenabled"].split(",")
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")
        target_type = SpiderFootHelpers.targetTypeFromString(record.target)
        if target_type is None:
            continue
        new_scan_id = SpiderFootHelpers.genScanInstanceId()
        try:
            # Try Celery first (v5.4.0+)
            celery_dispatched = False
            try:
                from spiderfoot.celery_app import is_celery_available
                if is_celery_available():
                    from spiderfoot.tasks.scan import run_scan
                    run_scan.apply_async(
                        kwargs={
                            "scan_name": record.name,
                            "scan_target": record.target,
                            "module_list": modlist,
                            "type_list": [],
                            "global_opts": cfg,
                        },
                        task_id=new_scan_id,
                        queue="scan",
                    )
                    celery_dispatched = True
            except Exception:
                pass

            if not celery_dispatched:
                p = mp.Process(
                    target=startSpiderFootScanner,
                    args=(None, record.name, new_scan_id, record.target, target_type, modlist, cfg),
                )
                p.daemon = True
                p.start()
        except Exception as e:
            continue
        while dbh.scanInstanceGet(new_scan_id) is None:
            time.sleep(1)
        new_scan_ids.append(new_scan_id)

    return {"new_scan_ids": new_scan_ids, "message": f"{len(new_scan_ids)} scans rerun started"}


# -----------------------------------------------------------------------
# Bulk operations
# -----------------------------------------------------------------------

class BulkScanRequest(BaseModel):
    """Request body for bulk scan operations."""
    scan_ids: list[str] = Field(..., description="List of scan IDs to operate on", min_length=1, max_length=100)


@router.post("/scans/bulk/stop")
async def bulk_stop_scans(
    request: BulkScanRequest,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Stop multiple scans in one request."""
    results = {"stopped": [], "not_found": [], "already_finished": [], "errors": []}
    for scan_id in request.scan_ids:
        try:
            record = svc.get_scan(scan_id)
            if record is None:
                results["not_found"].append(scan_id)
                continue
            status = record.status if hasattr(record, "status") else str(record)
            if status in (DB_STATUS_FINISHED, DB_STATUS_ABORTED, DB_STATUS_ERROR_FAILED):
                results["already_finished"].append(scan_id)
                continue
            svc.stop_scan(scan_id)
            results["stopped"].append(scan_id)
            if _hooks:
                try:
                    _hooks.on_stopped(scan_id)
                except Exception as e:
                    log.debug("on_stopped hook failed for scan %s: %s", scan_id, e)
        except Exception as e:
            log.error("Bulk stop failed for %s: %s", scan_id, e)
            results["errors"].append({"scan_id": scan_id, "error": str(e)})
    return {
        **results,
        "summary": {
            "stopped": len(results["stopped"]),
            "not_found": len(results["not_found"]),
            "already_finished": len(results["already_finished"]),
            "errors": len(results["errors"]),
        },
    }


@router.post("/scans/bulk/delete")
async def bulk_delete_scans(
    request: BulkScanRequest,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Delete multiple scans in one request."""
    results = {"deleted": [], "not_found": [], "errors": []}
    for scan_id in request.scan_ids:
        try:
            record = svc.get_scan(scan_id)
            if record is None:
                results["not_found"].append(scan_id)
                continue
            svc.delete_scan(scan_id)
            results["deleted"].append(scan_id)
            if _hooks:
                try:
                    _hooks.on_deleted(scan_id)
                except Exception as e:
                    log.debug("on_deleted hook failed for scan %s: %s", scan_id, e)
        except Exception as e:
            log.error("Bulk delete failed for %s: %s", scan_id, e)
            results["errors"].append({"scan_id": scan_id, "error": str(e)})
    return {
        **results,
        "summary": {
            "deleted": len(results["deleted"]),
            "not_found": len(results["not_found"]),
            "errors": len(results["errors"]),
        },
    }


@router.post("/scans/bulk/archive")
async def bulk_archive_scans(
    request: BulkScanRequest,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Archive multiple scans in one request."""
    results = {"archived": [], "not_found": [], "errors": []}
    for scan_id in request.scan_ids:
        try:
            record = svc.get_scan(scan_id)
            if record is None:
                results["not_found"].append(scan_id)
                continue
            svc.archive(scan_id)
            results["archived"].append(scan_id)
            if _hooks:
                try:
                    _hooks.on_archived(scan_id)
                except Exception as e:
                    log.debug("on_archived hook failed for scan %s: %s", scan_id, e)
        except Exception as e:
            log.error("Bulk archive failed for %s: %s", scan_id, e)
            results["errors"].append({"scan_id": scan_id, "error": str(e)})
    return {
        **results,
        "summary": {
            "archived": len(results["archived"]),
            "not_found": len(results["not_found"]),
            "errors": len(results["errors"]),
        },
    }


# -----------------------------------------------------------------------
# Recurring scan schedules
# -----------------------------------------------------------------------

@router.get("/scans/schedules")
async def list_schedules(api_key: str = optional_auth_dep) -> dict:
    """List all recurring scan schedules."""
    try:
        from spiderfoot.recurring_schedule import get_recurring_scheduler
        scheduler = get_recurring_scheduler()
        schedules = scheduler.list_all()
        return {
            "schedules": [s.to_dict() for s in schedules],
            "count": len(schedules),
            "stats": scheduler.stats(),
        }
    except Exception as e:
        log.error("Failed to list schedules: %s", e)
        raise HTTPException(status_code=500, detail="Failed to list schedules") from e


@router.post("/scans/schedules", status_code=201)
async def create_schedule(
    body: ScheduleCreateRequest,
    api_key: str = api_key_dep,
) -> dict:
    """Create a new recurring scan schedule."""
    if not body.interval_minutes and not body.run_at:
        raise HTTPException(
            status_code=422,
            detail="Either interval_minutes (>0) or run_at must be provided",
        )
    try:
        from spiderfoot.recurring_schedule import get_recurring_scheduler
        scheduler = get_recurring_scheduler()
        schedule = scheduler.add_schedule(
            name=body.name,
            target=body.target,
            interval_minutes=body.interval_minutes,
            run_at=body.run_at,
            modules=body.modules,
            type_filter=body.type_filter,
            max_runs=body.max_runs,
            description=body.description,
            tags=body.tags,
        )
        return {
            "schedule_id": schedule.schedule_id,
            "message": "Schedule created",
            **schedule.to_dict(),
        }
    except Exception as e:
        log.error("Failed to create schedule: %s", e)
        raise HTTPException(status_code=500, detail="Failed to create schedule") from e


@router.get("/scans/schedules/{schedule_id}")
async def get_schedule(schedule_id: str, api_key: str = optional_auth_dep) -> dict:
    """Get details of a specific scan schedule."""
    from spiderfoot.recurring_schedule import get_recurring_scheduler
    scheduler = get_recurring_scheduler()
    s = scheduler.get(schedule_id)
    if s is None:
        raise HTTPException(status_code=404, detail="Schedule not found")
    return s.to_dict()


@router.delete("/scans/schedules/{schedule_id}")
async def delete_schedule(schedule_id: str, api_key: str = api_key_dep) -> dict:
    """Delete a scan schedule."""
    from spiderfoot.recurring_schedule import get_recurring_scheduler
    scheduler = get_recurring_scheduler()
    if not scheduler.remove(schedule_id):
        raise HTTPException(status_code=404, detail="Schedule not found")
    return {"schedule_id": schedule_id, "message": "Schedule deleted"}


@router.post("/scans/schedules/{schedule_id}/pause")
async def pause_schedule(schedule_id: str, api_key: str = api_key_dep) -> dict:
    """Pause a recurring scan schedule."""
    from spiderfoot.recurring_schedule import get_recurring_scheduler
    scheduler = get_recurring_scheduler()
    if not scheduler.pause(schedule_id):
        raise HTTPException(status_code=404, detail="Schedule not found or not active")
    return {"schedule_id": schedule_id, "status": "paused"}


@router.post("/scans/schedules/{schedule_id}/resume")
async def resume_schedule(schedule_id: str, api_key: str = api_key_dep) -> dict:
    """Resume a paused scan schedule."""
    from spiderfoot.recurring_schedule import get_recurring_scheduler
    scheduler = get_recurring_scheduler()
    if not scheduler.resume(schedule_id):
        raise HTTPException(status_code=404, detail="Schedule not found or not paused")
    s = scheduler.get(schedule_id)
    return {
        "schedule_id": schedule_id,
        "status": "active",
        "next_run_at": s.next_run_at if s else None,
    }


# -----------------------------------------------------------------------
# Parameterized (CRUD + lifecycle) routes
# -----------------------------------------------------------------------


@router.get("/scans")
async def list_scans(
    params: PaginationParams = Depends(),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """List all scans with pagination."""
    try:
        records = svc.list_scans()
        dicts = [r.to_dict() for r in records]
        return paginate(dicts, params)
    except Exception as e:
        log.error("Failed to list scans: %s", e)
        raise HTTPException(status_code=500, detail="Failed to list scans") from e


@router.get("/scans/search")
async def search_scans(
    target: str | None = Query(None, description="Filter by target (substring match)"),
    status: str | None = Query(None, description="Filter by status (RUNNING, FINISHED, ABORTED, etc.)"),
    tag: str | None = Query(None, description="Filter by tag"),
    started_after: str | None = Query(None, description="Scans started after this ISO timestamp"),
    started_before: str | None = Query(None, description="Scans started before this ISO timestamp"),
    module: str | None = Query(None, description="Scans that used this module"),
    sort_by: str | None = Query("started", description="Sort field: started, target, status"),
    sort_order: str | None = Query("desc", description="Sort order: asc or desc"),
    limit: int = Query(50, ge=1, le=500, description="Max results"),
    offset: int = Query(0, ge=0, description="Offset for pagination"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Search and filter scans by target, status, tags, date range, and module.

    Returns matching scans with facets (status counts, target summary).
    """
    try:
        records = svc.list_scans()
        dicts = [r.to_dict() for r in records]

        # Apply filters
        filtered = []
        for s in dicts:
            # Target substring match
            if target:
                scan_target = str(s.get("target", s.get("name", "")))
                if target.lower() not in scan_target.lower():
                    continue

            # Status match
            if status:
                scan_status = str(s.get("status", "")).upper()
                if scan_status != status.upper():
                    continue

            # Tag match
            if tag:
                meta = s.get("metadata", s.get("meta", {})) or {}
                scan_tags = meta.get(_TAGS_KEY, [])
                if tag.lower() not in [t.lower() for t in scan_tags]:
                    continue

            # Date range
            scan_started = s.get("started", s.get("created", 0))
            if started_after:
                try:
                    from datetime import datetime
                    threshold = datetime.fromisoformat(started_after.replace("Z", "+00:00")).timestamp()
                    if isinstance(scan_started, (int, float)) and scan_started < threshold:
                        continue
                except (ValueError, TypeError):
                    pass

            if started_before:
                try:
                    from datetime import datetime
                    threshold = datetime.fromisoformat(started_before.replace("Z", "+00:00")).timestamp()
                    if isinstance(scan_started, (int, float)) and scan_started > threshold:
                        continue
                except (ValueError, TypeError):
                    pass

            # Module filter
            if module:
                scan_modules = s.get("modules", s.get("module_list", []))
                if isinstance(scan_modules, str):
                    scan_modules = [m.strip() for m in scan_modules.split(",")]
                if module not in scan_modules:
                    continue

            filtered.append(s)

        # Build facets before pagination
        status_counts: dict = {}
        for s in filtered:
            st = str(s.get("status", "UNKNOWN")).upper()
            status_counts[st] = status_counts.get(st, 0) + 1

        # Sort
        sort_key = sort_by if sort_by in ("started", "target", "status") else "started"
        reverse = sort_order != "asc"
        filtered.sort(
            key=lambda s: s.get(sort_key, s.get("created", "")),
            reverse=reverse,
        )

        # Pagination
        total = len(filtered)
        page = filtered[offset : offset + limit]

        return {
            "total": total,
            "offset": offset,
            "limit": limit,
            "scans": page,
            "facets": {
                "status": status_counts,
            },
        }
    except Exception as e:
        log.error("Failed to search scans: %s", e)
        raise HTTPException(status_code=500, detail="Failed to search scans") from e


@router.post("/scans", status_code=201, response_model=ScanCreateResponse)
async def create_scan(
    scan_request: ScanRequest,
    background_tasks: BackgroundTasks,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanCreateResponse:
    """Create and start a new scan."""
    try:
        config = get_app_config()
        scan_id = SpiderFootHelpers.genScanInstanceId()
        target_type = SpiderFootHelpers.targetTypeFromString(scan_request.target)
        if not target_type:
            raise HTTPException(status_code=422, detail="Invalid target")

        sf = SpiderFoot(config.get_config())
        sf_config = config.get_config()

        # If an engine profile is specified, load it and merge settings
        if scan_request.engine:
            try:
                from spiderfoot.scan_engine import ScanEngineLoader
                loader = ScanEngineLoader()
                engine = loader.load(scan_request.engine)
                modules = engine.get_enabled_modules()
                sf_config = engine.to_sf_config(sf_config)
                log.info("Using scan engine '%s' for scan %s", scan_request.engine, scan_id)
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Invalid engine: {e}")
        elif scan_request.profile:
            # Load modules from a scan profile (e.g. "tools-only", "quick-recon")
            try:
                from spiderfoot.scan.scan_profile import get_profile_manager
                pm = get_profile_manager()
                profile = pm.get(scan_request.profile)
                if not profile:
                    raise HTTPException(
                        status_code=422,
                        detail=f"Unknown scan profile: '{scan_request.profile}'",
                    )
                modules = list(profile.include_modules)
                log.info(
                    "Using scan profile '%s' (%d modules) for scan %s",
                    scan_request.profile, len(modules), scan_id,
                )
            except HTTPException:
                raise
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Invalid profile: {e}")
        else:
            all_modules = sf.modulesProducing(["*"])
            modules = scan_request.modules if scan_request.modules else all_modules

        if not modules:
            modules = ["sfp__stor_db"]

        # Always ensure storage modules are included
        # sfp__stor_db watches all events and stores results to the database
        # It won't appear in modulesProducing() since it doesn't produce events
        for storage_mod in ["sfp__stor_db"]:
            if storage_mod not in modules:
                modules.append(storage_mod)

        try:
            svc.create_scan(scan_id, scan_request.name, scan_request.target)
        except Exception as e:
            log.error("Failed to create scan instance: %s", e)
            raise HTTPException(
                status_code=500, detail="Unable to create scan instance in database"
            ) from e

        background_tasks.add_task(
            start_scan_background,
            scan_id,
            scan_request.name,
            scan_request.target,
            target_type,
            modules,
            scan_request.type_filter,
            sf_config,
        )
        return ScanCreateResponse(
            id=scan_id,
            name=scan_request.name,
            target=scan_request.target,
            status=DB_STATUS_STARTING,
            message="Scan created and starting",
        )
    except HTTPException:
        raise
    except Exception as e:
        log.error("Failed to create scan: %s", e)
        raise HTTPException(status_code=500, detail="Failed to create scan") from e
    finally:
        if _hooks:
            try:
                _hooks.on_created(scan_id, scan_request.name, scan_request.target)
            except Exception as e:
                log.debug("on_created hook failed for scan %s: %s", scan_id, e)


@router.get("/scans/{scan_id}")
async def get_scan(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan details."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    result = record.to_dict()
    try:
        result["state_machine"] = svc.get_scan_state(scan_id)
    except Exception as e:
        log.debug("Failed to retrieve state machine for scan %s: %s", scan_id, e)
    return result


# -----------------------------------------------------------------------
# Scan result / event data endpoints (consumed by WebUI)
# -----------------------------------------------------------------------


@router.get("/scans/{scan_id}/events")
async def get_scan_events(
    scan_id: str,
    event_type: str = Query(None, description="Filter by event type"),
    filter_fp: bool = Query(False, description="Filter false positives"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan result events as JSON."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_events(scan_id, event_type=event_type or "ALL", filter_fp=filter_fp)
    events = []
    for row in rows:
        # DB returns: (generated, data, module, hash, type, source_event_hash, confidence, visibility, risk)
        events.append({
            "generated": row[0] if len(row) > 0 else 0,
            "data": row[1] if len(row) > 1 else "",
            "module": row[2] if len(row) > 2 else "",
            "hash": row[3] if len(row) > 3 else "",
            "type": row[4] if len(row) > 4 else "",
            "source_event_hash": row[5] if len(row) > 5 else "ROOT",
            "confidence": row[6] if len(row) > 6 else 100,
            "visibility": row[7] if len(row) > 7 else 100,
            "risk": row[8] if len(row) > 8 else 0,
        })
    return {"events": events, "total": len(events)}


@router.get("/scans/{scan_id}/summary")
async def get_scan_summary(
    scan_id: str,
    by: str = Query("type", description="Group by: type, module, entity"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan result summary grouped by type/module/entity."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_result_summary(scan_id, by)
    # DB returns: (type/module/data, event_descr, last_in, total, utotal)
    summary = []
    for row in rows:
        summary.append({
            "key": row[0] if len(row) > 0 else "",
            "description": row[1] if len(row) > 1 else "",
            "last_in": row[2] if len(row) > 2 else 0,
            "total": row[3] if len(row) > 3 else 0,
            "unique_total": row[4] if len(row) > 4 else 0,
        })
    # Also produce a simple dict form for the WebUI client
    summary_dict = {row[0]: row[3] for row in rows if len(row) > 3}
    return {"summary": summary_dict, "details": summary, "total_types": len(summary)}


@router.get("/scans/{scan_id}/logs")
async def get_scan_logs(
    scan_id: str,
    limit: int = Query(None, description="Max log entries"),
    offset: int = Query(0, description="Offset (fromRowId)"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan log entries."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_scan_logs(scan_id, limit=limit, from_row_id=offset)
    # DB returns: (generated, component, type, message, rowid)
    logs = []
    for row in rows:
        logs.append({
            "generated": row[0] if len(row) > 0 else 0,
            "component": row[1] if len(row) > 1 else "",
            "type": row[2] if len(row) > 2 else "",
            "message": row[3] if len(row) > 3 else "",
            "rowid": row[4] if len(row) > 4 else 0,
        })
    return {"logs": logs, "total": len(logs)}


@router.get("/scans/{scan_id}/history")
async def get_scan_history(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan result history (counts over time)."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_result_history(scan_id)
    # DB returns: (hourmin, type, count)
    history = []
    for row in rows:
        history.append({
            "time": row[0] if len(row) > 0 else "",
            "type": row[1] if len(row) > 1 else "",
            "count": row[2] if len(row) > 2 else 0,
        })
    return {"history": history}


@router.get("/scans/{scan_id}/events/unique")
async def get_scan_events_unique(
    scan_id: str,
    event_type: str = Query("ALL", description="Filter by event type"),
    filterfp: bool = Query(False, description="Filter false positives"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get unique scan result events."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_unique_events(scan_id, event_type, filterfp)
    # DB returns: (data, type, count)
    events = []
    for row in rows:
        events.append({
            "data": row[0] if len(row) > 0 else "",
            "type": row[1] if len(row) > 1 else "",
            "count": row[2] if len(row) > 2 else 0,
        })
    return {"events": events, "total": len(events)}


@router.get("/scans/{scan_id}/correlations")
async def get_scan_correlations(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan correlations."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    rows = svc.get_correlations(scan_id)
    correlations = []
    for row in rows:
        if isinstance(row, dict):
            correlations.append(row)
        elif isinstance(row, (list, tuple)):
            # DB returns: id, title, rule_id, rule_risk, rule_name,
            #             rule_descr, rule_logic, event_count
            correlations.append({
                "id": row[0] if len(row) > 0 else "",
                "title": row[1] if len(row) > 1 else "",
                "rule_id": row[2] if len(row) > 2 else "",
                "rule_risk": row[3] if len(row) > 3 else "",
                "rule_name": row[4] if len(row) > 4 else "",
                "rule_descr": row[5] if len(row) > 5 else "",
                "rule_logic": row[6] if len(row) > 6 else "",
                "event_count": row[7] if len(row) > 7 else 0,
            })
    return {"correlations": correlations, "total": len(correlations)}


@router.get("/scans/{scan_id}/correlations/summary")
async def get_scan_correlation_summary(
    scan_id: str,
    by: str = Query("risk", description="Group by: rule or risk"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get scan correlation summary."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    try:
        rows = svc.get_correlation_summary(scan_id, by)
        summary = []
        for row in rows:
            if isinstance(row, dict):
                summary.append(row)
            elif isinstance(row, (list, tuple)):
                if by == "risk":
                    summary.append({
                        "risk": row[0] if len(row) > 0 else "",
                        "total": row[1] if len(row) > 1 else 0,
                    })
                else:
                    summary.append({
                        "rule_id": row[0] if len(row) > 0 else "",
                        "rule_name": row[1] if len(row) > 1 else "",
                        "risk": row[2] if len(row) > 2 else "",
                        "description": row[3] if len(row) > 3 else "",
                        "total": row[4] if len(row) > 4 else 0,
                    })
        return {"summary": summary, "total": len(summary)}
    except Exception as e:
        log.debug("Correlation summary unavailable for %s: %s", scan_id, e)
        return {"summary": [], "total": 0}


@router.post("/scans/{scan_id}/correlations/run")
async def run_scan_correlations(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Run correlation rules against a scan's results.

    Loads all YAML correlation rules from the ``correlations/`` directory,
    executes them via ``RuleExecutor``, and stores matched correlations
    in ``tbl_scan_correlation_results``.
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    try:
        import os
        from spiderfoot.correlation.rule_executor import RuleExecutor

        # Load rules
        rules_dir = os.environ.get("SF_CORRELATION_RULES_DIR", "correlations")
        rules: list[dict] = []
        if os.path.isdir(rules_dir):
            try:
                from spiderfoot.correlation.rule_loader import RuleLoader
                loader = RuleLoader(rules_dir)
                rules = loader.load_rules()
            except ImportError:
                rules = SpiderFootHelpers.loadCorrelationRulesRaw(rules_dir) or []

        if not rules:
            return {"message": "No correlation rules found", "results": 0}

        # Get a DB handle for the executor
        dbh = svc._dbh if hasattr(svc, '_dbh') else None
        if dbh is None:
            try:
                from spiderfoot import SpiderFootDb
                db_path = os.environ.get("SF_DATABASE", "spiderfoot.db")
                dbh = SpiderFootDb({"__database": db_path}, init=True)
            except Exception as exc:
                log.error("Cannot get DB handle for correlations: %s", exc)
                raise HTTPException(status_code=500, detail="DB unavailable") from exc

        executor = RuleExecutor(dbh, rules, scan_ids=[scan_id])
        raw_results = executor.run()

        total = sum(
            rule_results.get("correlations_created", 0)
            for rule_results in raw_results.values()
            if isinstance(rule_results, dict)
        )

        return {
            "message": f"Correlation complete for scan {scan_id}",
            "rules_evaluated": len(rules),
            "results": total,
        }

    except HTTPException:
        raise
    except Exception as e:
        log.error("Correlation run failed for %s: %s", scan_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Correlation failed: {e}") from e


@router.delete("/scans/{scan_id}", response_model=ScanDeleteResponse)
async def delete_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanDeleteResponse:
    """Delete a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    svc.delete_scan(scan_id)
    if _hooks:
        try:
            _hooks.on_deleted(scan_id)
        except Exception as e:
            log.debug("on_deleted hook failed for scan %s: %s", scan_id, e)
    return ScanDeleteResponse()


@router.delete("/scans/{scan_id}/full", response_model=MessageResponse)
async def delete_scan_full(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> MessageResponse:
    """Delete a scan and all related data."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    svc.delete_scan_full(scan_id)
    return MessageResponse(message="Scan and all related data deleted successfully")


@router.post("/scans/{scan_id}/stop", response_model=ScanStopResponse)
async def stop_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanStopResponse:
    """Stop a running scan with state-machine validation."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    try:
        new_status = svc.stop_scan(scan_id)
        if _hooks:
            try:
                _hooks.on_aborted(scan_id, reason="API stop request")
            except Exception as e:
                log.debug("on_aborted hook failed for scan %s: %s", scan_id, e)
        return ScanStopResponse(message="Scan stopped successfully", status=new_status)
    except ScanServiceError as e:
        raise HTTPException(status_code=409, detail=str(e)) from e


@router.post("/scans/{scan_id}/retry")
async def retry_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Retry a failed or aborted scan by creating a new scan with the same configuration.

    The original scan is preserved for reference.  A new scan is created
    with the same target, modules, and configuration, and optionally
    started immediately.
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    scan_dict = record.to_dict()
    status = str(scan_dict.get("status", "")).upper()

    # Only allow retry of non-running scans
    if status in (DB_STATUS_RUNNING, DB_STATUS_STARTING, DB_STATUS_STARTED):
        raise HTTPException(
            status_code=409,
            detail=f"Cannot retry a scan in '{status}' state — stop it first",
        )

    try:
        from spiderfoot import SpiderFootHelpers

        # Get the original scan config
        dbh = svc._dbh if hasattr(svc, '_dbh') else None
        scan_config = {}
        original_target = scan_dict.get("target", scan_dict.get("name", ""))
        original_modules = scan_dict.get("modules", [])

        if dbh and hasattr(dbh, 'scanConfigGet'):
            try:
                scan_config = dbh.scanConfigGet(scan_id) or {}
            except Exception as e:
                log.debug("Failed to retrieve scan config for %s: %s", scan_id, e)

        new_scan_id = SpiderFootHelpers.genScanInstanceId()

        # Create the new scan entry
        if dbh and hasattr(dbh, 'scanInstanceCreate'):
            scan_name = f"{scan_dict.get('name', original_target)} (Retry)"
            dbh.scanInstanceCreate(new_scan_id, scan_name, original_target)
            if scan_config:
                dbh.scanConfigSet(new_scan_id, scan_config)

        # Copy metadata (including tags, annotations) from original
        try:
            original_meta = svc.get_metadata(scan_id) or {}
            if original_meta:
                retry_meta = dict(original_meta)
                retry_meta["_retry_of"] = scan_id
                retry_meta["_retry_reason"] = status
                svc.set_metadata(new_scan_id, retry_meta)
        except Exception as e:
            log.debug("Failed to copy metadata from scan %s to %s: %s", scan_id, new_scan_id, e)

        if _hooks:
            try:
                _hooks.on_created(new_scan_id, original_target)
            except Exception as e:
                log.debug("on_created hook failed for retry scan %s: %s", new_scan_id, e)

        return {
            "original_scan_id": scan_id,
            "original_status": status,
            "new_scan_id": new_scan_id,
            "target": original_target,
            "message": "New scan created from original configuration. Start it with POST /scans/{id}/start.",
        }
    except HTTPException:
        raise
    except Exception as e:
        log.error("Failed to retry scan %s: %s", scan_id, e)
        raise HTTPException(status_code=500, detail="Failed to retry scan") from e


# -----------------------------------------------------------------------
# Export endpoints
# -----------------------------------------------------------------------


@router.get("/scans/{scan_id}/events/export")
async def export_scan_event_results(
    scan_id: str,
    event_type: str = None,
    filetype: str = "csv",
    dialect: str = "excel",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> StreamingResponse:
    """Export scan event result data as CSV or Excel."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    data = svc.get_events(scan_id, event_type if event_type else "ALL")
    rows = []
    for row in data:
        if row[4] == "ROOT":
            continue
        lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
        datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
        rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

    headings = ["Updated", "Type", "Module", "Source", "F/P", "Data"]

    if filetype.lower() in ("xlsx", "excel"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EventResults"
        ws.append(headings)
        for r in rows:
            ws.append(r)
        with BytesIO() as f:
            wb.save(f)
            f.seek(0)
            return StreamingResponse(
                iter([f.read()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-eventresults.xlsx",
                    "Pragma": "no-cache",
                },
            )

    if filetype.lower() == "csv":
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(headings)
        for r in rows:
            parser.writerow(r)
        fileobj.seek(0)
        return StreamingResponse(
            iter([fileobj.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-eventresults.csv",
                "Pragma": "no-cache",
            },
        )

    if filetype.lower() == "json":
        json_rows = []
        for r in rows:
            json_rows.append({
                "updated": r[0],
                "type": r[1],
                "module": r[2],
                "source": r[3],
                "false_positive": r[4],
                "data": r[5],
            })
        payload = json.dumps({
            "scan_id": scan_id,
            "total": len(json_rows),
            "events": json_rows,
        }, indent=2)
        return Response(
            content=payload,
            media_type="application/json",
            headers={
                "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-eventresults.json",
                "Pragma": "no-cache",
            },
        )

    if filetype.lower() == "gexf":
        # Redirect to the viz endpoint which handles GEXF generation
        return await export_scan_viz(scan_id, gexf="1", api_key=api_key, svc=svc)

    raise HTTPException(status_code=400, detail="Invalid export filetype.")


@router.get("/scans/{scan_id}/search/export")
async def export_scan_search_results(
    scan_id: str,
    event_type: str = None,
    value: str = None,
    filetype: str = "csv",
    dialect: str = "excel",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> StreamingResponse:
    """Export search result data as CSV or Excel."""
    data = []
    for row in svc.search_events(scan_id, event_type=event_type or "", value=value or ""):
        if len(row) < 12 or row[10] == "ROOT":
            continue
        datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
        data.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])

    headings = ["Updated", "Type", "Module", "Source", "F/P", "Data"]

    if not data:
        raise HTTPException(status_code=404, detail="No search results found for this scan")

    if filetype.lower() in ("xlsx", "excel"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SearchResults"
        ws.append(headings)
        for r in data:
            ws.append(r)
        with BytesIO() as f:
            wb.save(f)
            f.seek(0)
            return StreamingResponse(
                iter([f.read()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-searchresults.xlsx",
                    "Pragma": "no-cache",
                },
            )

    if filetype.lower() == "csv":
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(headings)
        for r in data:
            parser.writerow(r)
        fileobj.seek(0)
        return StreamingResponse(
            iter([fileobj.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-searchresults.csv",
                "Pragma": "no-cache",
            },
        )

    raise HTTPException(status_code=400, detail="Invalid export filetype.")


@router.get("/scans/{scan_id}/viz")
async def export_scan_viz(
    scan_id: str,
    gexf: str = "0",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> Response:
    """Export entities from scan results for visualising (GEXF/graph)."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    events = svc.get_events(scan_id, filter_fp=True)

    if not events:
        payload = json.dumps({"nodes": [], "edges": []})
        return Response(payload, media_type="application/json")

    root = record.target

    # Build event-type → category lookup
    event_type_categories: dict[str, str] = {}
    try:
        dbh = svc._ensure_dbh()
        et_list = dbh.eventTypes()
        if et_list:
            for et in et_list:
                if isinstance(et, (list, tuple)):
                    if len(et) >= 4:
                        event_type_categories[et[1]] = et[3]
                    elif len(et) >= 2:
                        event_type_categories[et[0]] = et[1] if len(et) > 1 else "DATA"
                elif isinstance(et, dict):
                    event_type_categories[et.get("event", "")] = et.get("event_type", "DATA")
    except Exception:
        pass

    # Build hash → data-value mapping for parent resolution
    hash_to_data: dict[str, str] = {}
    for row in events:
        row = list(row) if isinstance(row, tuple) else row
        event_hash = row[3] if len(row) > 3 else ""
        event_data = row[1] if len(row) > 1 else ""
        if event_hash:
            hash_to_data[event_hash] = event_data

    # Transform 9-col DB rows → 15-col format expected by buildGraphData
    extended_data = []
    for row in events:
        row = list(row) if isinstance(row, tuple) else row
        generated = row[0] if len(row) > 0 else 0
        data_val = row[1] if len(row) > 1 else ""
        module = row[2] if len(row) > 2 else ""
        event_hash = row[3] if len(row) > 3 else ""
        event_type = row[4] if len(row) > 4 else ""
        source_hash = row[5] if len(row) > 5 else "ROOT"
        confidence = row[6] if len(row) > 6 else 100
        visibility = row[7] if len(row) > 7 else 100
        risk = row[8] if len(row) > 8 else 0

        source_data = hash_to_data.get(source_hash, "ROOT") if source_hash != "ROOT" else root
        category = event_type_categories.get(event_type, "DATA")

        extended_data.append((
            generated,    # 0
            data_val,     # 1 — entity value
            source_data,  # 2 — parent entity value
            module,       # 3
            event_type,   # 4
            source_hash,  # 5
            confidence,   # 6
            visibility,   # 7
            event_hash,   # 8 — event ID
            0,            # 9 — false_positive
            risk,         # 10
            category,     # 11 — ENTITY/INTERNAL/DESCRIPTOR/DATA
            "",           # 12
            "",           # 13
            "",           # 14
        ))

    if gexf == "0":
        graph_json = SpiderFootHelpers.buildGraphJson([root], extended_data)
        return Response(graph_json, media_type="application/json")

    fname = f"{record.name}-SpiderFoot.gexf" if record.name else "SpiderFoot.gexf"
    gexf_data = SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", extended_data)
    return Response(
        gexf_data,
        media_type="application/gexf",
        headers={"Content-Disposition": f"attachment; filename={fname}", "Pragma": "no-cache"},
    )


@router.get("/scans/{scan_id}/logs/export")
async def export_scan_logs(
    scan_id: str,
    dialect: str = "excel",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> StreamingResponse:
    """Export scan logs as CSV."""
    try:
        data = svc.get_scan_logs(scan_id)
    except Exception as e:
        log.debug("Failed to export scan logs for %s: %s", scan_id, e)
        raise HTTPException(status_code=404, detail="Scan ID not found") from e

    if not data:
        raise HTTPException(status_code=404, detail="No scan logs found")

    fileobj = StringIO()
    parser = csv.writer(fileobj, dialect=dialect)
    parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
    for row in data:
        parser.writerow([str(x) for x in row])
    fileobj.seek(0)
    return StreamingResponse(
        iter([fileobj.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}.log.csv",
            "Pragma": "no-cache",
        },
    )


@router.get("/scans/{scan_id}/timeline")
async def get_scan_timeline(
    scan_id: str,
    limit: int = Query(200, ge=1, le=5000, description="Max timeline entries"),
    event_type: str | None = Query(None, description="Filter by event type"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Get a chronological timeline of events for a scan.

    Returns events ordered by timestamp with module attribution,
    useful for understanding scan progression and discovery order.
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    try:
        events = svc.get_events(scan_id) or []

        timeline = []
        for row in events:
            if isinstance(row, dict):
                et = row.get("type", row.get("eventType", ""))
                data = str(row.get("data", ""))
                module = row.get("module", "")
                ts = row.get("generated", row.get("lastSeen", 0))
            elif isinstance(row, (list, tuple)):
                ts = row[0] if len(row) > 0 else 0
                data = str(row[1]) if len(row) > 1 else ""
                module = str(row[3]) if len(row) > 3 else ""
                et = str(row[4]) if len(row) > 4 else ""
            else:
                continue

            if et == "ROOT":
                continue
            if event_type and et != event_type:
                continue

            # Clean up SFURL markers
            data = data.replace("<SFURL>", "").replace("</SFURL>", "")

            timeline.append({
                "timestamp": ts,
                "event_type": et,
                "data": data[:500],  # Truncate large data
                "module": module,
            })

        # Sort by timestamp (ascending = chronological)
        timeline.sort(key=lambda e: e["timestamp"])
        timeline = timeline[:limit]

        # Summary statistics
        type_counts = {}
        module_counts = {}
        for entry in timeline:
            type_counts[entry["event_type"]] = type_counts.get(entry["event_type"], 0) + 1
            module_counts[entry["module"]] = module_counts.get(entry["module"], 0) + 1

        return {
            "scan_id": scan_id,
            "total_events": len(events),
            "timeline_entries": len(timeline),
            "timeline": timeline,
            "summary": {
                "event_types": dict(sorted(type_counts.items(), key=lambda x: -x[1])),
                "modules": dict(sorted(module_counts.items(), key=lambda x: -x[1])),
                "first_event": timeline[0]["timestamp"] if timeline else None,
                "last_event": timeline[-1]["timestamp"] if timeline else None,
            },
        }
    except Exception as e:
        log.error("Failed to get scan timeline for %s: %s", scan_id, e)
        raise HTTPException(status_code=500, detail="Failed to build scan timeline") from e


@router.get("/scans/{scan_id}/dedup")
async def detect_duplicate_events(
    scan_id: str,
    threshold: int = Query(2, ge=2, le=100, description="Min occurrences to flag as duplicate"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Detect duplicate events in scan results.

    Finds events with identical (event_type, data) pairs across modules,
    helping identify redundant data collection and module overlap.
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    try:
        events = svc.get_events(scan_id) or []

        # Build fingerprint → occurrences map
        fingerprints: dict = {}  # (event_type, data_hash) → list of sources
        for row in events:
            if isinstance(row, dict):
                et = row.get("type", row.get("eventType", ""))
                data = str(row.get("data", ""))
                module = row.get("module", "")
            elif isinstance(row, (list, tuple)):
                data = str(row[1]) if len(row) > 1 else ""
                module = str(row[3]) if len(row) > 3 else ""
                et = str(row[4]) if len(row) > 4 else ""
            else:
                continue

            if et == "ROOT":
                continue

            key = (et, data[:500])
            if key not in fingerprints:
                fingerprints[key] = []
            fingerprints[key].append(module)

        # Filter to duplicates only
        duplicates = []
        for (et, data_preview), modules in fingerprints.items():
            if len(modules) >= threshold:
                duplicates.append({
                    "event_type": et,
                    "data": data_preview[:200],
                    "occurrences": len(modules),
                    "modules": list(set(modules)),
                })

        duplicates.sort(key=lambda d: -d["occurrences"])

        total_events = len(events)
        total_dup_events = sum(d["occurrences"] for d in duplicates)

        return {
            "scan_id": scan_id,
            "total_events": total_events,
            "unique_fingerprints": len(fingerprints),
            "duplicate_groups": len(duplicates),
            "duplicate_event_count": total_dup_events,
            "dedup_ratio": round(1 - len(fingerprints) / max(total_events, 1), 4),
            "duplicates": duplicates[:200],
        }
    except Exception as e:
        log.error("Failed to detect duplicates for %s: %s", scan_id, e)
        raise HTTPException(status_code=500, detail="Failed to detect duplicates") from e


@router.get("/scans/{scan_id}/correlations/export")
async def export_scan_correlations(
    scan_id: str,
    filetype: str = "csv",
    dialect: str = "excel",
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> StreamingResponse:
    """Export scan correlation data as CSV or Excel."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    try:
        data = svc.get_correlations(scan_id)
    except Exception as e:
        raise HTTPException(status_code=404, detail="Scan ID not found")

    headings = ["ID", "Title", "Rule ID", "Risk", "Rule Name", "Description", "Logic", "Events"]

    if filetype.lower() in ("xlsx", "excel"):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Correlations"
        ws.append(headings)
        for row in data:
            if isinstance(row, (list, tuple)):
                ws.append(list(row))
            elif isinstance(row, dict):
                ws.append([row.get(h.lower().replace(" ", "_"), "") for h in headings])
        with BytesIO() as f:
            wb.save(f)
            content = f.getvalue()
        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-correlations.xlsx",
                "Pragma": "no-cache",
            },
        )

    if filetype.lower() == "csv":
        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(headings)
        for row in data:
            if isinstance(row, dict):
                parser.writerow([str(row.get(h.lower().replace(" ", "_"), "")) for h in headings])
            else:
                parser.writerow([str(x) for x in row])
        fileobj.seek(0)
        return StreamingResponse(
            iter([fileobj.getvalue()]),
            media_type="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=SpiderFoot-{scan_id}-correlations.csv",
                "Pragma": "no-cache",
            },
        )

    raise HTTPException(status_code=400, detail="Invalid export filetype.")


# -----------------------------------------------------------------------
# Lifecycle (rerun, clone)
# -----------------------------------------------------------------------


@router.get("/scans/{scan_id}/options")
async def get_scan_options(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Return configuration used for the specified scan."""
    config = get_app_config()
    ret = svc.get_scan_options(scan_id, config.get_config())
    return ret


@router.post("/scans/{scan_id}/rerun", response_model=ScanRerunResponse)
async def rerun_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanRerunResponse:
    """Rerun a scan."""
    config = get_app_config()
    cfg = deepcopy(config.get_config())
    dbh = svc.dbh

    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Invalid scan ID.")
    if not record.target:
        raise HTTPException(status_code=400, detail=f"Scan {scan_id} has no target defined.")

    scanconfig = dbh.scanConfigGet(scan_id)
    if not scanconfig:
        raise HTTPException(status_code=400, detail=f"Error loading config from scan: {scan_id}")

    modlist = scanconfig["_modulesenabled"].split(",")
    if "sfp__stor_stdout" in modlist:
        modlist.remove("sfp__stor_stdout")

    target_type = SpiderFootHelpers.targetTypeFromString(record.target)
    if not target_type:
        target_type = SpiderFootHelpers.targetTypeFromString(f'"{record.target}"')
    if not target_type:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot determine target type for scan rerun. Target '{record.target}' is not recognized.",
        )

    scantarget = record.target
    if target_type not in ("HUMAN_NAME", "BITCOIN_ADDRESS"):
        scantarget = scantarget.lower()

    new_scan_id = SpiderFootHelpers.genScanInstanceId()
    try:
        # Try Celery first (v5.4.0+)
        celery_dispatched = False
        try:
            from spiderfoot.celery_app import is_celery_available
            if is_celery_available():
                from spiderfoot.tasks.scan import run_scan
                run_scan.apply_async(
                    kwargs={
                        "scan_name": record.name,
                        "scan_target": scantarget,
                        "module_list": modlist,
                        "type_list": [],
                        "global_opts": cfg,
                    },
                    task_id=new_scan_id,
                    queue="scan",
                )
                celery_dispatched = True
                log.info("Scan rerun %s dispatched to Celery worker", new_scan_id)
        except Exception as e:
            log.warning("Celery dispatch failed for rerun %s, using mp.Process: %s", new_scan_id, e)

        if not celery_dispatched:
            p = mp.Process(
                target=startSpiderFootScanner,
                args=(None, record.name, new_scan_id, scantarget, target_type, modlist, cfg),
            )
            p.daemon = True
            p.start()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scan [{new_scan_id}] failed: {e}")

    while dbh.scanInstanceGet(new_scan_id) is None:
        time.sleep(1)

    return ScanRerunResponse(new_scan_id=new_scan_id)


@router.post("/scans/{scan_id}/clone", response_model=ScanCloneResponse)
async def clone_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanCloneResponse:
    """Clone a scan configuration (without running it)."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Invalid scan ID.")

    dbh = svc.dbh
    scanconfig = dbh.scanConfigGet(scan_id)
    if not scanconfig:
        raise HTTPException(status_code=400, detail=f"Error loading config from scan: {scan_id}")

    modlist = scanconfig["_modulesenabled"].split(",")
    if "sfp__stor_stdout" in modlist:
        modlist.remove("sfp__stor_stdout")

    target_type = SpiderFootHelpers.targetTypeFromString(record.target)
    if not target_type:
        target_type = SpiderFootHelpers.targetTypeFromString(f'"{record.target}"')
    if not target_type:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot determine target type for scan clone. Target '{record.target}' is not recognized.",
        )

    scantarget = record.target
    if target_type not in ("HUMAN_NAME", "BITCOIN_ADDRESS"):
        scantarget = scantarget.lower()

    new_scan_id = SpiderFootHelpers.genScanInstanceId()
    try:
        dbh.scanInstanceCreate(new_scan_id, f"{record.name} (Clone)", scantarget)
        dbh.scanConfigSet(new_scan_id, scanconfig)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Scan [{new_scan_id}] clone failed: {e}") from e

    return ScanCloneResponse(new_scan_id=new_scan_id)


# -----------------------------------------------------------------------
# Event annotations (per-result notes/comments)
# -----------------------------------------------------------------------

_ANNOTATIONS_KEY = "_annotations"


@router.get("/scans/{scan_id}/annotations")
async def list_event_annotations(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """List all event annotations for a scan.

    Annotations are operator notes attached to individual scan result
    events, identified by result ID.
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    meta = svc.get_metadata(scan_id) or {}
    annotations = meta.get(_ANNOTATIONS_KEY, {})
    return {
        "scan_id": scan_id,
        "total": len(annotations),
        "annotations": annotations,
    }


@router.put("/scans/{scan_id}/annotations/{result_id}")
async def set_event_annotation(
    scan_id: str,
    result_id: str,
    note: str = Body(..., embed=True, max_length=2000),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Add or update an annotation on a specific scan result event.

    Args:
        result_id: The unique ID of the scan result event.
        note: The annotation text (max 2000 chars).
    """
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    import time as _time
    meta = svc.get_metadata(scan_id) or {}
    annotations = meta.get(_ANNOTATIONS_KEY, {})

    is_new = result_id not in annotations
    annotations[result_id] = {
        "note": note.strip(),
        "updated_at": _time.time(),
    }
    meta[_ANNOTATIONS_KEY] = annotations
    svc.set_metadata(scan_id, meta)

    return {
        "scan_id": scan_id,
        "result_id": result_id,
        "action": "created" if is_new else "updated",
        "annotation": annotations[result_id],
    }


@router.delete("/scans/{scan_id}/annotations/{result_id}")
async def delete_event_annotation(
    scan_id: str,
    result_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Remove an annotation from a scan result event."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")

    meta = svc.get_metadata(scan_id) or {}
    annotations = meta.get(_ANNOTATIONS_KEY, {})

    if result_id not in annotations:
        raise HTTPException(status_code=404, detail="Annotation not found")

    del annotations[result_id]
    meta[_ANNOTATIONS_KEY] = annotations
    svc.set_metadata(scan_id, meta)

    return {"scan_id": scan_id, "result_id": result_id, "deleted": True}


# -----------------------------------------------------------------------
# Results management
# -----------------------------------------------------------------------


@router.post("/scans/{scan_id}/results/falsepositive")
async def set_results_false_positive(
    scan_id: str,
    resultids: list[str] = Body(...),
    fp: str = Body(...),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Set a batch of results as false positive or not."""
    if fp not in ("0", "1"):
        raise HTTPException(status_code=400, detail="No FP flag set or not set correctly.")

    try:
        return svc.set_false_positive(scan_id, resultids, fp)
    except ScanServiceError as e:
        raise HTTPException(status_code=404, detail=str(e)) from e


@router.post("/scans/{scan_id}/clear", response_model=MessageResponse)
async def clear_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> MessageResponse:
    """Remove all results/events for a scan, keeping the scan entry."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    try:
        svc.clear_results(scan_id)
        return MessageResponse(message="Scan results cleared (scan entry retained)")
    except Exception as e:
        log.error("Failed to clear scan %s: %s", scan_id, e)
        raise HTTPException(status_code=500, detail="Failed to clear scan results") from e


# -----------------------------------------------------------------------
# Metadata / notes / archive
# -----------------------------------------------------------------------


@router.get("/scans/{scan_id}/metadata", response_model=ScanMetadataResponse)
async def get_scan_metadata(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanMetadataResponse:
    """Get scan metadata."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    return ScanMetadataResponse(metadata=svc.get_metadata(scan_id))


@router.patch("/scans/{scan_id}/metadata")
async def update_scan_metadata(
    scan_id: str,
    metadata: dict = Body(...),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Update scan metadata (key/value pairs)."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    if not isinstance(metadata, dict):
        raise HTTPException(status_code=422, detail="Metadata must be a dictionary")
    svc.set_metadata(scan_id, metadata)
    return {"success": True, "metadata": metadata}


@router.get("/scans/{scan_id}/notes", response_model=ScanNotesResponse)
async def get_scan_notes(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanNotesResponse:
    """Get scan notes/comments."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    return ScanNotesResponse(notes=svc.get_notes(scan_id))


@router.patch("/scans/{scan_id}/notes")
async def update_scan_notes(
    scan_id: str,
    notes: str = Body(...),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Update scan notes/comments."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    svc.set_notes(scan_id, notes)
    return {"success": True, "notes": notes}


@router.post("/scans/{scan_id}/archive", response_model=MessageResponse)
async def archive_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> MessageResponse:
    """Archive a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    svc.archive(scan_id)
    if _hooks:
        try:
            _hooks.on_archived(scan_id)
        except Exception as e:
            logging.debug("Hook callback error", exc_info=True)
    return MessageResponse(message="Scan archived")


@router.post("/scans/{scan_id}/unarchive", response_model=MessageResponse)
async def unarchive_scan(
    scan_id: str,
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> MessageResponse:
    """Unarchive a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    svc.unarchive(scan_id)
    if _hooks:
        try:
            _hooks.on_unarchived(scan_id)
        except Exception as e:
            logging.debug("Hook callback error", exc_info=True)
    return MessageResponse(message="Scan unarchived")


# -----------------------------------------------------------------------
# Scan tags / labels
# -----------------------------------------------------------------------

_TAGS_KEY = "_tags"


@router.get("/scans/{scan_id}/tags", response_model=ScanTagsResponse)
async def get_scan_tags(
    scan_id: str,
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanTagsResponse:
    """Get all tags for a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    meta = svc.get_metadata(scan_id) or {}
    tags = meta.get(_TAGS_KEY, [])
    return ScanTagsResponse(scan_id=scan_id, tags=tags)


@router.put("/scans/{scan_id}/tags", response_model=ScanTagsResponse)
async def set_scan_tags(
    scan_id: str,
    tags: list[str] = Body(..., description="Complete list of tags to set"),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanTagsResponse:
    """Replace all tags for a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    # Normalize: strip, lowercase, deduplicate, remove empty
    clean = sorted({t.strip().lower() for t in tags if t.strip()})
    if len(clean) > 50:
        raise HTTPException(status_code=422, detail="Maximum 50 tags per scan")
    meta = svc.get_metadata(scan_id) or {}
    meta[_TAGS_KEY] = clean
    svc.set_metadata(scan_id, meta)
    return ScanTagsResponse(scan_id=scan_id, tags=clean, message="Tags updated")


@router.post("/scans/{scan_id}/tags", response_model=ScanTagsResponse)
async def add_scan_tags(
    scan_id: str,
    tags: list[str] = Body(..., description="Tags to add"),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanTagsResponse:
    """Add tags to a scan (merges with existing)."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    meta = svc.get_metadata(scan_id) or {}
    existing = set(meta.get(_TAGS_KEY, []))
    new_tags = {t.strip().lower() for t in tags if t.strip()}
    merged = sorted(existing | new_tags)
    if len(merged) > 50:
        raise HTTPException(status_code=422, detail="Maximum 50 tags per scan")
    meta[_TAGS_KEY] = merged
    svc.set_metadata(scan_id, meta)
    added = sorted(new_tags - existing)
    return ScanTagsResponse(
        scan_id=scan_id, tags=merged,
        message=f"Added {len(added)} tag(s)" if added else "No new tags added",
    )


@router.delete("/scans/{scan_id}/tags", response_model=ScanTagsResponse)
async def remove_scan_tags(
    scan_id: str,
    tags: list[str] = Body(..., description="Tags to remove"),
    api_key: str = api_key_dep,
    svc: ScanService = Depends(get_scan_service),
) -> ScanTagsResponse:
    """Remove specific tags from a scan."""
    record = svc.get_scan(scan_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Scan not found")
    meta = svc.get_metadata(scan_id) or {}
    existing = set(meta.get(_TAGS_KEY, []))
    to_remove = {t.strip().lower() for t in tags if t.strip()}
    remaining = sorted(existing - to_remove)
    removed = sorted(existing & to_remove)
    meta[_TAGS_KEY] = remaining
    svc.set_metadata(scan_id, meta)
    return ScanTagsResponse(
        scan_id=scan_id, tags=remaining,
        message=f"Removed {len(removed)} tag(s)" if removed else "No matching tags found",
    )
from spiderfoot.scan.scan_state_map import (
    DB_STATUS_ABORTED,
    DB_STATUS_ERROR_FAILED,
    DB_STATUS_FINISHED,
    DB_STATUS_RUNNING,
    DB_STATUS_STARTED,
    DB_STATUS_STARTING,
)


# -----------------------------------------------------------------------
# Scan comparison
# -----------------------------------------------------------------------


@router.get("/scans/compare")
async def compare_scans(
    scan_a: str = Query(..., description="First scan ID"),
    scan_b: str = Query(..., description="Second scan ID"),
    api_key: str = optional_auth_dep,
    svc: ScanService = Depends(get_scan_service),
) -> dict:
    """Compare two scans and return the diff of their findings.

    Returns event types and data that are:
    - **only_in_a**: found in scan A but not scan B
    - **only_in_b**: found in scan B but not scan A
    - **common**: found in both scans

    Useful for tracking changes between re-scans of the same target.
    """
    # Verify both scans exist
    rec_a = svc.get_scan(scan_a)
    rec_b = svc.get_scan(scan_b)
    if rec_a is None:
        raise HTTPException(status_code=404, detail=f"Scan '{scan_a}' not found")
    if rec_b is None:
        raise HTTPException(status_code=404, detail=f"Scan '{scan_b}' not found")

    try:
        # Get events for both scans
        events_a = svc.get_events(scan_a) or []
        events_b = svc.get_events(scan_b) or []

        # Build sets of (eventType, data) tuples for comparison
        def _event_key(e):
            if isinstance(e, dict):
                return (e.get("type", e.get("eventType", "")), str(e.get("data", "")))
            return (getattr(e, "eventType", ""), str(getattr(e, "data", "")))

        set_a = {_event_key(e) for e in events_a}
        set_b = {_event_key(e) for e in events_b}

        only_a = set_a - set_b
        only_b = set_b - set_a
        common = set_a & set_b

        # Group by event type
        def _group_by_type(items):
            grouped = {}
            for etype, data in items:
                grouped.setdefault(etype, []).append(data)
            return grouped

        # Type-level summary
        types_a = {t for t, _ in set_a}
        types_b = {t for t, _ in set_b}

        return {
            "scan_a": {"id": scan_a, "total_events": len(events_a)},
            "scan_b": {"id": scan_b, "total_events": len(events_b)},
            "summary": {
                "only_in_a": len(only_a),
                "only_in_b": len(only_b),
                "common": len(common),
                "new_event_types_in_b": sorted(types_b - types_a),
                "removed_event_types_in_b": sorted(types_a - types_b),
            },
            "diff": {
                "only_in_a": _group_by_type(sorted(only_a)[:500]),
                "only_in_b": _group_by_type(sorted(only_b)[:500]),
            },
        }
    except Exception as e:
        log.error("Failed to compare scans %s vs %s: %s", scan_a, scan_b, e)
        raise HTTPException(status_code=500, detail="Scan comparison failed") from e
